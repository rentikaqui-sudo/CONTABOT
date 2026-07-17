"""
ContaBot — Flask API multi-contador
Cada contador se registra con email/password y solo ve sus propias empresas.
Data layer: Supabase (service_role key)
"""

import os, json, functools, re, sys, logging, time, secrets, threading
from datetime import date, timedelta, datetime, timezone
from pathlib import Path
from urllib.parse import urlencode
import urllib.request as _urllib_req
import bcrypt
from flask import Flask, jsonify, send_from_directory, request, session, redirect
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from werkzeug.middleware.proxy_fix import ProxyFix
from dotenv import load_dotenv
from supabase import create_client, Client

logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s %(message)s")

# ── Configuración ─────────────────────────────────────────────────────────────

BASE_DIR     = os.path.dirname(os.path.dirname(__file__))
UI_DIR       = os.path.join(BASE_DIR, "ui")
SCRIPTS_DIR  = os.path.join(BASE_DIR, "scripts")
FACTURAS_DIR = Path(BASE_DIR) / "data" / "facturas_recibidas"

load_dotenv(os.path.join(BASE_DIR, ".env"))

sys.path.insert(0, SCRIPTS_DIR)
from extractor import extraer_xml, extraer_pdf, descomprimir_zip, detectar_empresa, detectar_o_crear_empresa, guardar_empresa_pendiente, guardar_factura, determinar_flujo, COLORES, ICONOS
from telegram_notif import notificar_factura, notificar_empresa_desconocida
from calendario import todas_las_obligaciones, obligaciones_proximas

SUPABASE_URL = os.environ["SUPABASE_URL"]
SUPABASE_KEY = os.environ["SUPABASE_SERVICE_KEY"]

sb: Client = create_client(SUPABASE_URL, SUPABASE_KEY)


def _migrate_contadores():
    """Agrega columnas telegram si no existen (idempotente)."""
    try:
        sb.table("contadores").update({"telegram_chat_id": None, "telegram_token": None}).eq("id", 0).execute()
    except Exception:
        pass  # Columnas no existen — requieren migración manual en Supabase SQL Editor


_migrate_contadores()

app = Flask(__name__, static_folder=UI_DIR)
app.wsgi_app = ProxyFix(app.wsgi_app, x_proto=1, x_host=1)
_flask_secret = os.environ.get("FLASK_SECRET_KEY")
if not _flask_secret:
    raise RuntimeError("FLASK_SECRET_KEY no configurado. Generá uno con: python -c \"import secrets; print(secrets.token_hex(32))\"")
app.secret_key = _flask_secret
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=os.environ.get("FLASK_ENV") == "production",
    PERMANENT_SESSION_LIFETIME=timedelta(hours=8),
    MAX_CONTENT_LENGTH=20 * 1024 * 1024,
)

limiter = Limiter(get_remote_address, app=app, default_limits=[])


# ── Cifrado de tokens OAuth ───────────────────────────────────────────────────

def _get_fernet():
    key = os.environ.get("ENCRYPTION_KEY", "")
    if not key:
        return None
    try:
        from cryptography.fernet import Fernet
        return Fernet(key.encode() if isinstance(key, str) else key)
    except Exception:
        return None

def _encrypt_token(token: str) -> str:
    f = _get_fernet()
    if not f:
        return token
    return f.encrypt(token.encode()).decode()

def _decrypt_token(token: str) -> str:
    f = _get_fernet()
    if not f:
        return token
    try:
        return f.decrypt(token.encode()).decode()
    except Exception:
        return token  # plaintext fallback for tokens stored before encryption was enabled


# ── Cache en memoria (TTL 60s por contador) ───────────────────────────────────

_cache: dict = {}
_CACHE_TTL = 60

def _cache_get(key):
    if key in _cache:
        val, ts = _cache[key]
        if time.time() - ts < _CACHE_TTL:
            return val
        del _cache[key]
    return None

def _cache_set(key, val):
    _cache[key] = (val, time.time())

def _cache_invalidar(cid):
    for k in list(_cache.keys()):
        if k.startswith(f"{cid}:"):
            del _cache[k]


# ── Auth helpers ──────────────────────────────────────────────────────────────

def login_required(f):
    @functools.wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("contador_id"):
            return jsonify({"error": "No autorizado", "redirect": "/login"}), 401
        return f(*args, **kwargs)
    return decorated


def validate_empresa_ownership(eid):
    """Retorna respuesta 403 si la empresa no pertenece al contador en sesión, None si ok."""
    cid = session.get("contador_id")
    row = sb.table("empresas_clientes").select("id").eq("id", eid).eq("contador_id", cid).execute().data
    if not row:
        return jsonify({"error": "Sin acceso a esta empresa"}), 403
    return None


def get_user_empresa_ids():
    """Lista de empresa_ids del contador autenticado."""
    cid = session.get("contador_id")
    rows = sb.table("empresas_clientes").select("id").eq("contador_id", cid).execute().data
    return [r["id"] for r in rows]


def _calcular_estado_factura(fecha_vencimiento, dias_pago=None):
    """Calcula estado PENDIENTE/POR_VENCER/VENCIDA/PAGADA a partir de la fecha de vencimiento."""
    if dias_pago == 0:
        return "PAGADA"
    if not fecha_vencimiento:
        return "PENDIENTE"
    try:
        hoy = date.today()
        vto = date.fromisoformat(str(fecha_vencimiento)[:10])
        if vto < hoy:
            return f"VENCIDA ({(hoy - vto).days} dias)"
        if (vto - hoy).days <= 7:
            return "POR_VENCER"
        return "PENDIENTE"
    except Exception:
        return "PENDIENTE"


def _insertar_factura_gasto(empresa_id, datos, fuente="manual", archivo_url=None):
    """Único punto de inserción/actualización de facturas_gastos."""
    estado = _calcular_estado_factura(datos.get("fecha_vencimiento"))
    row = {
        "empresa_id":        empresa_id,
        "numero":            datos["numero"],
        "cufe":              datos.get("cufe", ""),
        "fecha":             datos.get("fecha", date.today().isoformat()),
        "fecha_vencimiento": datos.get("fecha_vencimiento"),
        "proveedor_nit":     datos.get("proveedor_nit", ""),
        "proveedor_nombre":  datos.get("proveedor_nombre", ""),
        "proveedor_ciudad":  datos.get("proveedor_ciudad", ""),
        "categoria":         datos.get("categoria", ""),
        "subtotal":          float(datos.get("subtotal", 0) or 0),
        "iva":               float(datos.get("iva", 0) or 0),
        "retefuente":        float(datos.get("retefuente", 0) or 0),
        "reteiva":           float(datos.get("reteiva", 0) or 0),
        "reteica":           float(datos.get("reteica", 0) or 0),
        "total_factura":     float(datos.get("total_factura", 0) or 0),
        "valor_neto":        float(datos.get("valor_neto", 0) or 0),
        "estado":            datos.get("estado", estado),
        "archivo_pdf":       archivo_url or datos.get("archivo_pdf", ""),
        "fuente":            fuente,
    }
    return sb.table("facturas_gastos").upsert(row, on_conflict="empresa_id,numero").execute()


@app.after_request
def security_headers(response):
    response.headers["X-Content-Type-Options"] = "nosniff"
    response.headers["X-Frame-Options"] = "DENY"
    response.headers["Referrer-Policy"] = "strict-origin-when-cross-origin"
    response.headers["Content-Security-Policy"] = (
        "default-src 'self'; "
        "script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; "
        "style-src 'self' 'unsafe-inline' https://fonts.googleapis.com; "
        "font-src 'self' https://fonts.gstatic.com; "
        "img-src 'self' data:; "
        "connect-src 'self';"
    )
    return response


# ── Páginas ──────────────────────────────────────────────────────────────────

@app.route("/health")
def health():
    try:
        supabase.table("contadores").select("id").limit(1).execute()
        return jsonify({"status": "ok", "db": "ok"})
    except Exception:
        return jsonify({"status": "ok", "db": "unreachable"}), 200


@app.route("/api/cron/recordar-tokens-gmail")
def cron_recordar_tokens_gmail():
    """Llamado por cron-job.org diariamente. Avisa cuando un token tiene 6+ días (vence a los 7)."""
    from datetime import datetime, timezone, timedelta
    WARN_AFTER = timedelta(days=6)
    try:
        tokens = sb.table("gmail_tokens").select(
            "empresa_id, email, activo, token_created_at, reminder_sent_at, "
            "empresas_clientes(razon_social, nit, contador_id)"
        ).eq("activo", True).execute().data

        avisados = 0
        for t in tokens:
            created_at = t.get("token_created_at")
            if not created_at:
                continue
            age = datetime.now(timezone.utc) - datetime.fromisoformat(created_at.replace("Z", "+00:00"))
            if age < WARN_AFTER:
                continue
            # No avisar dos veces por el mismo token
            reminder_sent = t.get("reminder_sent_at")
            if reminder_sent:
                reminded_at = datetime.fromisoformat(reminder_sent.replace("Z", "+00:00"))
                if reminded_at >= datetime.fromisoformat(created_at.replace("Z", "+00:00")):
                    continue
            empresa = t.get("empresas_clientes") or {}
            cid = empresa.get("contador_id")
            if not cid:
                continue
            chat_id = _tg_chat_id_for_contador(cid)
            if not chat_id:
                continue
            nombre = empresa.get("razon_social", f"empresa #{t['empresa_id']}")
            nit = empresa.get("nit", "")
            dias = age.days
            _tg_send_raw(chat_id,
                f"⚠️ *Token Gmail por vencer*\n\n"
                f"*Empresa:* {nombre}\n"
                f"*NIT:* {nit}\n"
                f"*Gmail:* {t.get('email','')}\n"
                f"*Antigüedad:* {dias} días (vence a los 7)\n\n"
                f"Entra a ContaBot → selecciona *{nombre}* → tab *Gmail* → botón *Reconectar Gmail*"
            )
            sb.table("gmail_tokens").update(
                {"reminder_sent_at": datetime.now(timezone.utc).isoformat()}
            ).eq("empresa_id", t["empresa_id"]).execute()
            avisados += 1

        return jsonify({"ok": True, "avisados": avisados, "revisados": len(tokens)})
    except Exception:
        logging.exception("Error en cron recordar-tokens-gmail")
        return jsonify({"ok": False, "error": "Error interno"}), 500


@app.route("/api/cron/chequear-tokens-gmail")
def cron_chequear_tokens_gmail():
    """Llamado por cron-job.org cada semana. Detecta tokens vencidos y avisa por Telegram."""
    secret = os.environ.get("CRON_SECRET", "")
    if secret and request.args.get("secret") != secret:
        return jsonify({"ok": False}), 401
    try:
        from gmail_facturas import get_gmail_from_supabase
        from google.auth.exceptions import RefreshError
        tokens = sb.table("gmail_tokens").select("empresa_id, email, activo").eq("activo", True).execute().data
        vencidos = []
        for t in tokens:
            eid = t["empresa_id"]
            try:
                get_gmail_from_supabase(eid)
            except RefreshError:
                sb.table("gmail_tokens").update({"activo": False}).eq("empresa_id", eid).execute()
                empresa = sb.table("empresas_clientes").select("razon_social, nit, contador_id").eq("id", eid).execute().data
                if not empresa:
                    continue
                e = empresa[0]
                vencidos.append(e)
                chat_id = _tg_chat_id_for_contador(e["contador_id"])
                _tg_send_raw(chat_id,
                    f"⚠️ *Token Gmail vencido*\n\n"
                    f"*Empresa:* {e['razon_social']}\n"
                    f"*NIT:* {e.get('nit','')}\n"
                    f"*Gmail:* {t.get('email','')}\n\n"
                    f"Entra a ContaBot → selecciona *{e['razon_social']}* → tab *Gmail* → botón *Reconectar Gmail*"
                )
            except Exception:
                pass
        return jsonify({"ok": True, "revisados": len(tokens), "vencidos": len(vencidos)})
    except Exception:
        logging.exception("Error en cron chequear-tokens-gmail")
        return jsonify({"ok": False, "error": "Error interno"}), 500

@app.route("/")
def index():
    if not session.get("contador_id"):
        return redirect("/login")
    return send_from_directory(UI_DIR, "index.html")

@app.route("/bienvenida")
def bienvenida():
    return send_from_directory(UI_DIR, "bienvenida.html")

@app.route("/login")
def login_page():
    if session.get("contador_id"):
        return redirect("/")
    return send_from_directory(UI_DIR, "login.html")

@app.route("/<path:f>")
def static_files(f):
    return send_from_directory(UI_DIR, f)


# ── Auth ─────────────────────────────────────────────────────────────────────

@app.route("/api/registro", methods=["POST"])
@limiter.limit("5 per minute")
def api_registro():
    data = request.get_json() or {}
    email   = (data.get("email") or "").strip().lower()
    password = data.get("password", "")
    nombre  = (data.get("nombre") or "").strip()
    codigo = (data.get("codigo") or "").strip()
    beta_code = os.environ.get("BETA_CODE", "")
    if beta_code and codigo != beta_code:
        return jsonify({"ok": False, "error": "Código de acceso inválido"}), 403
    if not email or not password or not nombre:
        return jsonify({"ok": False, "error": "Email, contraseña y nombre son requeridos"}), 400
    if len(password) < 8:
        return jsonify({"ok": False, "error": "La contraseña debe tener al menos 8 caracteres"}), 400
    existing = sb.table("contadores").select("id").eq("email", email).execute().data
    if existing:
        return jsonify({"ok": False, "error": "El email ya está registrado"}), 409
    pw_hash = bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()
    try:
        row = sb.table("contadores").insert({
            "email":          email,
            "password_hash":  pw_hash,
            "nombre":         nombre,
            "estudio_nombre": (data.get("estudio_nombre") or "").strip(),
            "tp_numero":      (data.get("tp_numero") or "").strip(),
            "telefono":       (data.get("telefono") or "").strip(),
        }).execute().data[0]
    except Exception:
        logging.exception("Error creando contador")
        return jsonify({"ok": False, "error": "Error interno al crear la cuenta"}), 500
    session.permanent = True
    session["contador_id"] = row["id"]
    session["email"]       = email
    session["nombre"]      = nombre
    return jsonify({"ok": True, "nombre": nombre})


@app.route("/api/login", methods=["POST"])
@limiter.limit("10 per minute")
def api_login():
    data     = request.get_json() or {}
    email    = (data.get("email") or "").strip().lower()
    password = data.get("password", "")
    if not email or not password:
        return jsonify({"ok": False, "error": "Email y contraseña requeridos"}), 400
    rows = sb.table("contadores").select("id,email,nombre,password_hash").eq("email", email).execute().data
    if not rows or not bcrypt.checkpw(password.encode(), rows[0]["password_hash"].encode()):
        return jsonify({"ok": False, "error": "Credenciales inválidas"}), 401
    u = rows[0]
    session.permanent = True
    session["contador_id"] = u["id"]
    session["email"]       = u["email"]
    session["nombre"]      = u["nombre"]
    return jsonify({"ok": True, "nombre": u["nombre"]})


@app.route("/api/me")
@login_required
def api_me():
    cid = session["contador_id"]
    row = sb.table("contadores").select("id,email,nombre,estudio_nombre,tp_numero,telefono").eq("id", cid).execute().data
    return jsonify(row[0] if row else {})


@app.route("/api/logout")
def api_logout():
    session.clear()
    return redirect("/login")


# ── Vista general: todas las empresas del contador ───────────────────────────

@app.route("/api/empresas")
@login_required
def empresas():
    cid    = session["contador_id"]
    cached = _cache_get(f"{cid}:empresas")
    if cached:
        return jsonify(cached)
    empresas_rows = sb.table("empresas_clientes").select("*").eq("contador_id", cid).order("id").execute().data
    resultado = []

    if not empresas_rows:
        return jsonify([])

    empresa_ids = [e["id"] for e in empresas_rows]
    # 2 queries en total en lugar de N+1
    ventas_all = sb.table("facturas_venta").select("empresa_id,valor_neto,estado,tipo_documento").in_("empresa_id", empresa_ids).execute().data
    gastos_all = sb.table("facturas_gastos").select("empresa_id,valor_neto,estado").in_("empresa_id", empresa_ids).execute().data

    ventas_by_empresa = {}
    gastos_by_empresa = {}
    for r in ventas_all:
        ventas_by_empresa.setdefault(r["empresa_id"], []).append(r)
    for r in gastos_all:
        gastos_by_empresa.setdefault(r["empresa_id"], []).append(r)

    for e in empresas_rows:
        eid = e["id"]
        ventas_rows = ventas_by_empresa.get(eid, [])
        gastos_rows = gastos_by_empresa.get(eid, [])

        def _signo_v(r): return -1 if r.get("tipo_documento") == "nota_credito" else 1
        v_n           = len(ventas_rows)
        v_neto        = sum(_signo_v(r) * (r["valor_neto"] or 0) for r in ventas_rows)
        v_por_cobrar  = sum(_signo_v(r) * (r["valor_neto"] or 0) for r in ventas_rows if r["estado"] == "PENDIENTE")
        v_vencido     = sum(_signo_v(r) * (r["valor_neto"] or 0) for r in ventas_rows if "VENCIDA" in str(r["estado"]).upper())
        v_cobrado     = sum(_signo_v(r) * (r["valor_neto"] or 0) for r in ventas_rows if r["estado"] == "PAGADA")
        v_n_vencidas  = sum(1 for r in ventas_rows if "VENCIDA" in str(r["estado"]).upper())
        v_n_por_vencer= sum(1 for r in ventas_rows if r["estado"] == "POR_VENCER")

        g_n          = len(gastos_rows)
        g_neto       = sum(r["valor_neto"] or 0 for r in gastos_rows)
        g_por_pagar  = sum(r["valor_neto"] or 0 for r in gastos_rows if r["estado"] == "PENDIENTE")
        g_pagado     = sum(r["valor_neto"] or 0 for r in gastos_rows if r["estado"] == "PAGADA")

        alertas  = v_n_vencidas + v_n_por_vencer
        semaforo = "verde"
        if v_n_vencidas >= 2:
            semaforo = "rojo"
        elif alertas > 0:
            semaforo = "amarillo"

        resultado.append({
            "id":           e["id"],
            "razon_social": e["razon_social"],
            "nit":          e["nit"],
            "sector":       e["sector"],
            "ciudad":       e["ciudad"],
            "contacto":     e["contacto"],
            "color":        e["color"],
            "icono":        e["icono"],
            "semaforo":     semaforo,
            "ventas": {
                "n":            v_n,
                "neto":         round(v_neto),
                "por_cobrar":   round(v_por_cobrar),
                "vencido":      round(v_vencido),
                "cobrado":      round(v_cobrado),
                "n_vencidas":   v_n_vencidas,
                "n_por_vencer": v_n_por_vencer,
            },
            "gastos": {
                "n":        g_n,
                "neto":     round(g_neto),
                "por_pagar":round(g_por_pagar),
                "pagado":   round(g_pagado),
            },
            "alertas": alertas,
        })

    _cache_set(f"{cid}:empresas", resultado)
    return jsonify(resultado)


@app.route("/api/empresas", methods=["POST"])
@login_required
def crear_empresa():
    body = request.get_json()
    nit          = (body.get("nit") or "").strip()
    razon_social = (body.get("razon_social") or "").strip()
    ciudad       = (body.get("ciudad") or "").strip()
    sector       = (body.get("sector") or "General").strip()
    contacto     = (body.get("contacto") or "").strip()

    if not nit or not razon_social:
        return jsonify({"ok": False, "error": "NIT y razón social son obligatorios"}), 400

    # Verificar duplicado
    existe = sb.table("empresas_clientes").select("id").eq("nit", nit).execute().data
    if existe:
        return jsonify({"ok": False, "error": f"Ya existe una empresa con NIT {nit}"}), 409

    cid     = session["contador_id"]
    count   = len(sb.table("empresas_clientes").select("id").eq("contador_id", cid).execute().data)
    color   = COLORES[count % len(COLORES)]
    icono   = ICONOS[count % len(ICONOS)]

    regimen = (body.get("regimen") or "Juridica").strip()

    row = sb.table("empresas_clientes").insert({
        "contador_id":  cid,
        "nit":          nit,
        "razon_social": razon_social,
        "ciudad":       ciudad,
        "sector":       sector,
        "contacto":     contacto,
        "color":        color,
        "icono":        icono,
        "regimen":      regimen,
    }).execute().data

    _cache_invalidar(cid)
    return jsonify({"ok": True, "empresa": row[0] if row else {}})


@app.route("/api/empresa/<int:eid>", methods=["PUT"])
@login_required
def editar_empresa(eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    body = request.get_json() or {}
    campos = {}
    for k in ("nit", "razon_social", "ciudad", "sector", "contacto", "regimen"):
        v = body.get(k)
        if v is not None:
            campos[k] = str(v).strip()
    if not campos:
        return jsonify({"ok": False, "error": "Nada que actualizar"}), 400
    row = sb.table("empresas_clientes").update(campos).eq("id", eid).execute().data
    return jsonify({"ok": True, "empresa": row[0] if row else {}})


# ── Resumen consolidado del contador ─────────────────────────────────────────

@app.route("/api/resumen")
@login_required
def resumen():
    cid    = session["contador_id"]
    cached = _cache_get(f"{cid}:resumen")
    if cached:
        return jsonify(cached)
    empresa_ids = get_user_empresa_ids()
    if not empresa_ids:
        return jsonify({"n_empresas": 0, "total_ventas": 0, "por_cobrar": 0,
                        "cartera_vencida": 0, "n_vencidas": 0, "n_por_vencer": 0,
                        "total_gastos": 0, "por_pagar": 0, "total_facturas": 0})
    n_empresas  = len(empresa_ids)
    ventas_rows = sb.table("facturas_venta").select("valor_neto,estado,tipo_documento").in_("empresa_id", empresa_ids).execute().data
    gastos_rows = sb.table("facturas_gastos").select("valor_neto,estado").in_("empresa_id", empresa_ids).execute().data

    def _sv(r): return -1 if r.get("tipo_documento") == "nota_credito" else 1
    v_neto        = sum(_sv(r) * (r["valor_neto"] or 0) for r in ventas_rows)
    v_por_cobrar  = sum(_sv(r) * (r["valor_neto"] or 0) for r in ventas_rows if r["estado"] == "PENDIENTE")
    v_vencido     = sum(_sv(r) * (r["valor_neto"] or 0) for r in ventas_rows if "VENCIDA" in str(r["estado"]).upper())
    v_n_vencidas  = sum(1 for r in ventas_rows if "VENCIDA" in str(r["estado"]).upper())
    v_n_por_vencer= sum(1 for r in ventas_rows if r["estado"] == "POR_VENCER")

    g_neto       = sum(r["valor_neto"] or 0 for r in gastos_rows)
    g_por_pagar  = sum(r["valor_neto"] or 0 for r in gastos_rows if r["estado"] == "PENDIENTE")

    payload = {
        "n_empresas":      n_empresas,
        "total_ventas":    round(v_neto),
        "por_cobrar":      round(v_por_cobrar),
        "cartera_vencida": round(v_vencido),
        "n_vencidas":      v_n_vencidas,
        "n_por_vencer":    v_n_por_vencer,
        "total_gastos":    round(g_neto),
        "por_pagar":       round(g_por_pagar),
        "total_facturas":  len(ventas_rows) + len(gastos_rows),
    }
    _cache_set(f"{cid}:resumen", payload)
    return jsonify(payload)


# ── Detalle de una empresa ────────────────────────────────────────────────────

@app.route("/api/empresa/<int:eid>/dashboard")
@login_required
def empresa_dashboard(eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    e_rows = sb.table("empresas_clientes").select("*").eq("id", eid).execute().data
    if not e_rows:
        return jsonify({"error": "Empresa no encontrada"}), 404
    e = e_rows[0]

    ventas_rows = sb.table("facturas_venta").select("valor_neto,estado,retefuente,reteiva,reteica,tipo_documento").eq("empresa_id", eid).execute().data
    gastos_rows = sb.table("facturas_gastos").select("valor_neto,estado,retefuente,reteiva,reteica").eq("empresa_id", eid).execute().data

    def _sd(r): return -1 if r.get("tipo_documento") == "nota_credito" else 1
    v_n         = len(ventas_rows)
    v_neto      = sum(_sd(r) * (r["valor_neto"] or 0) for r in ventas_rows)
    v_por_cobrar= sum(_sd(r) * (r["valor_neto"] or 0) for r in ventas_rows if r["estado"] == "PENDIENTE")
    v_vencido   = sum(_sd(r) * (r["valor_neto"] or 0) for r in ventas_rows if "VENCIDA" in str(r["estado"]).upper())
    v_cobrado   = sum(_sd(r) * (r["valor_neto"] or 0) for r in ventas_rows if r["estado"] == "PAGADA")

    g_n         = len(gastos_rows)
    g_neto      = sum(r["valor_neto"] or 0 for r in gastos_rows)
    g_por_pagar = sum(r["valor_neto"] or 0 for r in gastos_rows if r["estado"] == "PENDIENTE")
    g_pagado    = sum(r["valor_neto"] or 0 for r in gastos_rows if r["estado"] == "PAGADA")

    ret_v_rf = sum(_sd(r) * (r["retefuente"] or 0) for r in ventas_rows)
    ret_v_ri = sum(_sd(r) * (r["reteiva"]    or 0) for r in ventas_rows)
    ret_v_rc = sum(_sd(r) * (r["reteica"]    or 0) for r in ventas_rows)

    ret_g_rf = sum(r["retefuente"] or 0 for r in gastos_rows)
    ret_g_ri = sum(r["reteiva"]    or 0 for r in gastos_rows)
    ret_g_rc = sum(r["reteica"]    or 0 for r in gastos_rows)

    return jsonify({
        "empresa": e,
        "ventas": {
            "n":          v_n,
            "neto":       round(v_neto),
            "por_cobrar": round(v_por_cobrar),
            "vencido":    round(v_vencido),
            "cobrado":    round(v_cobrado),
        },
        "gastos": {
            "n":        g_n,
            "neto":     round(g_neto),
            "por_pagar":round(g_por_pagar),
            "pagado":   round(g_pagado),
        },
        "retenciones_ventas": {
            "retefuente": round(ret_v_rf),
            "reteiva":    round(ret_v_ri),
            "reteica":    round(ret_v_rc),
        },
        "retenciones_gastos": {
            "retefuente": round(ret_g_rf),
            "reteiva":    round(ret_g_ri),
            "reteica":    round(ret_g_rc),
        },
    })


@app.route("/api/empresa/<int:eid>/facturas/venta")
@login_required
def empresa_ventas(eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    per_page = min(int(request.args.get("per_page", 50)), 200)
    page     = max(int(request.args.get("page", 1)), 1)
    start    = (page - 1) * per_page
    result   = (
        sb.table("facturas_venta")
        .select("numero,fecha,fecha_vencimiento,cliente_nombre,cliente_ciudad,"
                "gran_contribuyente,subtotal,iva,retefuente,reteiva,reteica,"
                "total_factura,valor_neto,estado,tipo_documento,referencia_nc",
                count="exact")
        .eq("empresa_id", eid)
        .order("fecha", desc=True)
        .range(start, start + per_page - 1)
        .execute()
    )
    total = result.count or 0
    return jsonify({
        "data":     result.data,
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "pages":    max(1, (total + per_page - 1) // per_page),
    })


@app.route("/api/empresa/<int:eid>/facturas/gastos")
@login_required
def empresa_gastos(eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    per_page = min(int(request.args.get("per_page", 50)), 200)
    page     = max(int(request.args.get("page", 1)), 1)
    start    = (page - 1) * per_page
    result   = (
        sb.table("facturas_gastos")
        .select("numero,fecha,fecha_vencimiento,proveedor_nombre,proveedor_ciudad,"
                "categoria,subtotal,iva,retefuente,reteiva,reteica,"
                "total_factura,valor_neto,estado,cufe,archivo_pdf,tipo_documento,referencia_nc",
                count="exact")
        .eq("empresa_id", eid)
        .order("fecha", desc=True)
        .range(start, start + per_page - 1)
        .execute()
    )
    total = result.count or 0
    return jsonify({
        "data":     result.data,
        "total":    total,
        "page":     page,
        "per_page": per_page,
        "pages":    max(1, (total + per_page - 1) // per_page),
    })


@app.route("/api/factura/archivo")
@login_required
def factura_archivo():
    import re as _re
    path = request.args.get("path", "")
    if not path:
        return "No path", 400
    if path.startswith("http"):
        # Supabase Storage: .../facturas/{empresa_id}/...
        m = _re.search(r'/facturas/(\d+)/', path)
        if m:
            err = validate_empresa_ownership(int(m.group(1)))
            if err: return err
        return redirect(path)
    from flask import send_file
    allowed   = Path(FACTURAS_DIR).resolve()
    requested = Path(BASE_DIR, path).resolve()
    if not str(requested).startswith(str(allowed)):
        return "Acceso denegado", 403
    # Verificar que la empresa del path pertenece al contador en sesión
    try:
        eid = int(requested.relative_to(allowed).parts[0])
        err = validate_empresa_ownership(eid)
        if err: return err
    except (ValueError, IndexError):
        pass
    if requested.exists():
        return send_file(requested)
    return "Archivo no encontrado", 404


@app.route("/api/archivos")
@login_required
def api_archivos():
    """Devuelve todas las facturas agrupadas por empresa y mes, con URL de descarga."""
    empresa_ids = get_user_empresa_ids()
    if not empresa_ids:
        return jsonify([])
    empresas_rows = sb.table("empresas_clientes").select("id,razon_social,nit,color,icono").in_("id", empresa_ids).order("id").execute().data
    facturas_rows = sb.table("facturas_gastos").select(
        "id,numero,fecha,proveedor_nombre,total_factura,cufe,archivo_pdf,empresa_id"
    ).in_("empresa_id", empresa_ids).order("fecha", desc=True).limit(500).execute().data

    por_empresa = {}
    for e in empresas_rows:
        por_empresa[e["id"]] = {
            "empresa_id":   e["id"],
            "razon_social": e["razon_social"],
            "nit":          e["nit"],
            "color":        e.get("color", "#6366f1"),
            "icono":        e.get("icono", "🏢"),
            "meses":        {},
        }

    for f in facturas_rows:
        eid = f["empresa_id"]
        if eid not in por_empresa:
            continue
        mes = (f["fecha"] or "")[:7] or "sin-fecha"
        if mes not in por_empresa[eid]["meses"]:
            por_empresa[eid]["meses"][mes] = []
        por_empresa[eid]["meses"][mes].append({
            "numero":          f["numero"],
            "fecha":           f["fecha"],
            "proveedor":       f["proveedor_nombre"],
            "total":           f["total_factura"],
            "cufe":            f["cufe"],
            "archivo_url":     f["archivo_pdf"],
        })

    result = []
    for e in por_empresa.values():
        meses_list = []
        for mes_key in sorted(e["meses"].keys(), reverse=True):
            meses_list.append({"mes": mes_key, "facturas": e["meses"][mes_key]})
        e["meses"] = meses_list
        if meses_list:
            result.append(e)
    return jsonify(result)


@app.route("/api/admin/migrate-storage", methods=["POST"])
@login_required
def migrate_storage():
    """Migra archivos locales existentes a Supabase Storage."""
    from extractor import subir_a_storage
    empresa_ids = get_user_empresa_ids()
    if not empresa_ids:
        return jsonify({"migrados": 0, "ya_en_storage": 0, "errores": 0})
    rows = sb.table("facturas_gastos").select("id,numero,fecha,empresa_id,archivo_pdf").in_("empresa_id", empresa_ids).execute().data
    migrados, errores, ya_en_storage = 0, 0, 0
    for r in rows:
        path = r.get("archivo_pdf", "") or ""
        if path.startswith("http"):
            ya_en_storage += 1
            continue
        if not path or not os.path.exists(path):
            errores += 1
            continue
        url = subir_a_storage(path, r["empresa_id"], r["numero"], r["fecha"] or "", sb)
        if url:
            sb.table("facturas_gastos").update({"archivo_pdf": url}).eq("id", r["id"]).execute()
            migrados += 1
        else:
            errores += 1
    return jsonify({"migrados": migrados, "ya_en_storage": ya_en_storage, "errores": errores})


@app.route("/api/empresa/<int:eid>/alertas")
@login_required
def empresa_alertas(eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    e_rows = sb.table("empresas_clientes").select("razon_social").eq("id", eid).execute().data
    if not e_rows:
        return jsonify([])
    e = e_rows[0]

    rows = (
        sb.table("facturas_venta")
        .select("numero,cliente_nombre,cliente_ciudad,valor_neto,fecha_vencimiento,estado")
        .eq("empresa_id", eid)
        .order("fecha_vencimiento")
        .execute().data
    )
    # Filter: estado LIKE 'VENCIDA%' OR estado='POR_VENCER'
    rows = [r for r in rows if "VENCIDA" in str(r["estado"]).upper() or r["estado"] == "POR_VENCER"]

    result = []
    for d in rows:
        es_vencida = "VENCIDA" in d["estado"].upper()
        nombre_corto = d["cliente_nombre"].split()
        d["email_preview"] = (
            f"Estimado(a) {' '.join(nombre_corto[:2])},\n\n"
            f"Le recordamos que la factura {d['numero']} de {e['razon_social']}\n"
            f"por valor de ${d['valor_neto']:,.0f} COP se encuentra "
            f"{'VENCIDA' if es_vencida else 'proxima a vencer'}"
            f" (vencimiento: {d['fecha_vencimiento']}).\n\n"
            f"Le agradecemos gestionar el pago a la mayor brevedad posible.\n\n"
            f"Atentamente,\nEstudio Contable Aristizabal & Asociados\n"
            f"Tel: 315-890-4321 | info@contablearistizabal.com.co"
        ).replace(",", ".")
        result.append(d)
    return jsonify(result)


# ── Alertas globales (todas las empresas) ────────────────────────────────────

@app.route("/api/alertas/global")
@login_required
def alertas_global():
    empresa_ids = get_user_empresa_ids()
    if not empresa_ids:
        return jsonify([])

    # Fetch facturas with alert states — only for this contador's empresas
    fv_rows = (
        sb.table("facturas_venta")
        .select("numero,cliente_nombre,valor_neto,fecha_vencimiento,estado,empresa_id")
        .in_("empresa_id", empresa_ids)
        .order("fecha_vencimiento")
        .execute().data
    )
    fv_rows = [r for r in fv_rows if "VENCIDA" in str(r["estado"]).upper() or r["estado"] == "POR_VENCER"]

    # Fetch empresas for this contador only
    empresas_map = {
        e["id"]: e
        for e in sb.table("empresas_clientes").select("id,razon_social,color").in_("id", empresa_ids).execute().data
    }

    result = []
    for r in fv_rows:
        emp = empresas_map.get(r["empresa_id"], {})
        result.append({
            "numero":          r["numero"],
            "cliente_nombre":  r["cliente_nombre"],
            "valor_neto":      r["valor_neto"],
            "fecha_vencimiento":r["fecha_vencimiento"],
            "estado":          r["estado"],
            "empresa_nombre":  emp.get("razon_social", ""),
            "empresa_color":   emp.get("color", ""),
            "empresa_id":      r["empresa_id"],
        })
    return jsonify(result)


# ── Retenciones por cliente (para una empresa dada) ───────────────────────────

@app.route("/api/empresa/<int:eid>/retenciones-por-cliente")
@login_required
def retenciones_por_cliente(eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    rows = (
        sb.table("facturas_venta")
        .select("cliente_nit,cliente_nombre,cliente_ciudad,retefuente,reteiva,reteica,valor_neto,tipo_documento")
        .eq("empresa_id", eid)
        .execute().data
    )

    # GROUP BY cliente_nit — Notas Crédito restan (anulan o reducen la factura original)
    grupos = {}
    for r in rows:
        key = r["cliente_nit"]
        if key not in grupos:
            grupos[key] = {
                "cliente_nombre":  r["cliente_nombre"],
                "cliente_ciudad":  r["cliente_ciudad"],
                "n_facturas":      0,
                "retefuente":      0,
                "reteiva":         0,
                "reteica":         0,
                "valor_neto":      0,
            }
        g = grupos[key]
        signo = -1 if r.get("tipo_documento") == "nota_credito" else 1
        g["n_facturas"] += 1
        g["retefuente"] += signo * (r["retefuente"] or 0)
        g["reteiva"]    += signo * (r["reteiva"]    or 0)
        g["reteica"]    += signo * (r["reteica"]    or 0)
        g["valor_neto"] += signo * (r["valor_neto"] or 0)

    result = []
    for g in grupos.values():
        g["total_ret"] = g["retefuente"] + g["reteiva"] + g["reteica"]
        result.append(g)

    result.sort(key=lambda x: x["total_ret"], reverse=True)
    return jsonify(result)


# ── Factura manual (papel) ────────────────────────────────────────────────────

@app.route("/api/factura-manual", methods=["POST"])
@login_required
def factura_manual():
    data = request.get_json()

    err = validate_empresa_ownership(data.get("empresa_id"))
    if err: return err

    fecha     = data.get("fecha", date.today().isoformat())
    dias      = int(data.get("dias_pago", 30))
    fecha_vto = (date.fromisoformat(fecha) + timedelta(days=dias)).isoformat()

    hoy      = date.today()
    vto_date = date.fromisoformat(fecha_vto)
    if dias == 0:
        estado = "PAGADA"
    elif vto_date < hoy:
        estado = f"VENCIDA ({(hoy - vto_date).days} dias)"
    elif (vto_date - hoy).days <= 7:
        estado = "POR_VENCER"
    else:
        estado = "PENDIENTE"

    tipo    = data.get("tipo", "gasto")
    numero  = data.get("numero", "")
    eid_fm  = data.get("empresa_id")
    tabla_dup = "facturas_gastos" if tipo == "gasto" else "facturas_venta"
    es_duplicada = bool(sb.table(tabla_dup).select("id").eq("empresa_id", eid_fm).eq("numero", numero).execute().data)

    try:
        if tipo == "gasto":
            _insertar_factura_gasto(data["empresa_id"], {
                "numero":           data["numero"],
                "cufe":             "MANUAL-" + data["numero"],
                "fecha":            fecha,
                "fecha_vencimiento":fecha_vto,
                "proveedor_nit":    data.get("tercero_nit", ""),
                "proveedor_nombre": data.get("tercero_nombre", ""),
                "proveedor_ciudad": "Manual",
                "categoria":        data.get("categoria", "insumos"),
                "subtotal":         data["subtotal"],
                "iva":              data["iva"],
                "retefuente":       data["retefuente"],
                "reteiva":          data["reteiva"],
                "reteica":          data["reteica"],
                "total_factura":    data["total_factura"],
                "valor_neto":       data["valor_neto"],
                "estado":           estado,
            }, fuente="manual")
        else:
            sb.table("facturas_venta").upsert({
                "empresa_id":       data["empresa_id"],
                "numero":           data["numero"],
                "cufe":             "MANUAL-" + data["numero"],
                "fecha":            fecha,
                "fecha_vencimiento":fecha_vto,
                "cliente_nit":      data.get("tercero_nit", ""),
                "cliente_nombre":   data.get("tercero_nombre", ""),
                "cliente_ciudad":   "Manual",
                "gran_contribuyente":0,
                "subtotal":         data["subtotal"],
                "iva":              data["iva"],
                "retefuente":       data["retefuente"],
                "reteiva":          data["reteiva"],
                "reteica":          data["reteica"],
                "total_factura":    data["total_factura"],
                "valor_neto":       data["valor_neto"],
                "estado":           estado,
                "archivo_pdf":      "",
            }, on_conflict="empresa_id,numero").execute()
        _cache_invalidar(session["contador_id"])
        return jsonify({"ok": True, "duplicada": es_duplicada, "numero": numero})
    except Exception as ex:
        logging.exception('Error en endpoint')
        return jsonify({"ok": False, "error": "Error interno"}), 500


# ── Procesar imagen de factura (QR / OCR) ────────────────────────────────────

@app.route("/api/procesar-imagen", methods=["POST"])
@login_required
@limiter.limit("10 per minute")
def procesar_imagen():
    if "imagen" not in request.files:
        return jsonify({"ok": False, "error": "No se recibió imagen"}), 400

    archivo   = request.files["imagen"]
    img_bytes = archivo.read()
    nombre    = archivo.filename or "imagen"

    resultado = {"ok": True, "metodo": None, "datos": {}, "raw": "", "archivo": nombre}

    try:
        import cv2
        import numpy as np
    except ImportError:
        return jsonify({
            "ok": False,
            "error": "Falta opencv-python. Ejecute: pip install opencv-python pillow"
        }), 500

    nparr = np.frombuffer(img_bytes, np.uint8)
    img   = cv2.imdecode(nparr, cv2.IMREAD_COLOR)

    if img is None:
        return jsonify({"ok": False, "error": "No se pudo leer la imagen. Verifique el formato (JPG, PNG, etc.)"}), 400

    # ── Intentar leer QR ──────────────────────────────────────────────────────
    detector = cv2.QRCodeDetector()
    qr_data, bbox, _ = detector.detectAndDecode(img)

    if qr_data:
        resultado["metodo"] = "qr"
        resultado["raw"]    = qr_data
        resultado["datos"]  = _parse_dian_qr(qr_data)
        return jsonify(resultado)

    # ── Intentar OCR ─────────────────────────────────────────────────────────
    try:
        import pytesseract
        from PIL import Image as PILImage
        import io

        pil_img = PILImage.open(io.BytesIO(img_bytes))
        texto   = pytesseract.image_to_string(pil_img, lang="spa+eng")
        resultado["metodo"] = "ocr"
        resultado["raw"]    = texto
        resultado["datos"]  = _parse_ocr_texto(texto)
    except ImportError:
        resultado["metodo"] = "sin_ocr"
        resultado["datos"]  = {
            "confiabilidad": "N/A",
            "fuente": "No se encontró QR y pytesseract no está instalado",
        }
        resultado["mensaje"] = "No se detectó código QR en la imagen. Instale pytesseract + Tesseract para OCR."
    except Exception as e:
        resultado["metodo"] = "error_ocr"
        resultado["datos"]  = {"fuente": "Error en OCR", "confiabilidad": "N/A"}
        resultado["mensaje"] = f"No se pudo extraer texto: {str(e)}"

    return jsonify(resultado)


def _parse_dian_qr(qr_text):
    datos = {"tipo": "url_dian", "url": qr_text}

    cufe = re.search(r'documentkey=([a-f0-9]{96})', qr_text, re.IGNORECASE)
    if cufe:
        datos["cufe"]       = cufe.group(1)
        datos["cufe_corto"] = cufe.group(1)[:16] + "..." + cufe.group(1)[-8:]

    datos["fuente"]        = "QR de factura electrónica DIAN"
    datos["confiabilidad"] = "Alta — datos directamente del portal DIAN"
    return datos


def _parse_ocr_texto(texto):
    datos = {}

    m = re.search(r'NIT[:\s.]*([0-9.]{7,12}[-–]?\d)', texto, re.IGNORECASE)
    if m:
        datos["nit"] = m.group(1).strip()

    m = re.search(r'TOTAL[:\s]*\$?\s*([\d.,]+)', texto, re.IGNORECASE)
    if m:
        datos["total_texto"] = m.group(1)
        datos["total"]       = int(re.sub(r'[.,]', '', m.group(1))[:12])

    m = re.search(r'(\d{2}[/-]\d{2}[/-]\d{4}|\d{4}[/-]\d{2}[/-]\d{2})', texto)
    if m:
        datos["fecha"] = m.group(1)

    m = re.search(r'(?:factura|FV|FC|FE|N[°º])[:\s#-]*([A-Z0-9-]{3,20})', texto, re.IGNORECASE)
    if m:
        datos["numero"] = m.group(1).strip()

    m = re.search(r'CUFE[:\s]*([a-f0-9]{32,})', texto, re.IGNORECASE)
    if m:
        datos["cufe"] = m.group(1)[:96]

    datos["fuente"]        = "OCR — reconocimiento óptico de texto"
    datos["confiabilidad"] = "Media — verificar valores antes de guardar"
    return datos


# ── Informe mensual PDF ────────────────────────────────────────────────────────

@app.route("/api/empresa/<int:eid>/informe-pdf")
@login_required
@limiter.limit("10 per minute")
def informe_pdf(eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    from reportlab.lib.pagesizes import letter
    from reportlab.lib import colors
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer, HRFlowable)
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm
    from flask import make_response
    import io

    e_rows = sb.table("empresas_clientes").select("*").eq("id", eid).execute().data
    if not e_rows:
        return jsonify({"error": "Empresa no encontrada"}), 404
    e = e_rows[0]

    vs = (
        sb.table("facturas_venta")
        .select("numero,fecha,cliente_nombre,subtotal,iva,retefuente,reteiva,reteica,"
                "total_factura,valor_neto,estado")
        .eq("empresa_id", eid)
        .order("fecha")
        .execute().data
    )
    gs = (
        sb.table("facturas_gastos")
        .select("numero,fecha,proveedor_nombre,categoria,subtotal,iva,retefuente,reteiva,reteica,"
                "total_factura,valor_neto,estado")
        .eq("empresa_id", eid)
        .order("fecha")
        .execute().data
    )

    # Aggregate ventas totals
    vt = {
        "n":          len(vs),
        "total":      sum(r["total_factura"] or 0 for r in vs),
        "neto":       sum(r["valor_neto"]    or 0 for r in vs),
        "rf":         sum(r["retefuente"]    or 0 for r in vs),
        "ri":         sum(r["reteiva"]       or 0 for r in vs),
        "rc":         sum(r["reteica"]       or 0 for r in vs),
        "por_cobrar": sum(r["valor_neto"]    or 0 for r in vs if r["estado"] == "PENDIENTE"),
        "vencido":    sum(r["valor_neto"]    or 0 for r in vs if "VENCIDA" in str(r["estado"]).upper()),
    }
    gt = {
        "n":    len(gs),
        "total":sum(r["total_factura"] or 0 for r in gs),
        "neto": sum(r["valor_neto"]    or 0 for r in gs),
        "rf":   sum(r["retefuente"]    or 0 for r in gs),
        "ri":   sum(r["reteiva"]       or 0 for r in gs),
        "rc":   sum(r["reteica"]       or 0 for r in gs),
    }

    NAVY  = colors.HexColor('#1a2744')
    BLUE  = colors.HexColor('#2563eb')
    LGRAY = colors.HexColor('#f1f5f9')
    MGRAY = colors.HexColor('#94a3b8')
    GREEN = colors.HexColor('#059669')
    RED   = colors.HexColor('#dc2626')
    WHITE = colors.white
    DARK  = colors.HexColor('#1e293b')

    def fmt(n): return f"${int(n or 0):,}".replace(',', '.')

    def ps(text, size=9, bold=False, color=DARK, align=0):
        return Paragraph(str(text), ParagraphStyle('_',
            fontSize=size, textColor=color, leading=size*1.35,
            fontName='Helvetica-Bold' if bold else 'Helvetica', alignment=align))

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=letter,
                            rightMargin=1.8*cm, leftMargin=1.8*cm,
                            topMargin=1.8*cm, bottomMargin=1.8*cm)
    story = []

    # Header
    hdr = Table([[ps('ESTUDIO CONTABLE ARISTIZÁBAL & ASOCIADOS', 13, True, WHITE),
                  ps('INFORME FINANCIERO', 11, True, WHITE, 2)]],
                colWidths=['*', '*'])
    hdr.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),NAVY),
        ('TOPPADDING',(0,0),(-1,-1),12),('BOTTOMPADDING',(0,0),(-1,-1),12),
        ('LEFTPADDING',(0,0),(-1,-1),14),('RIGHTPADDING',(0,0),(-1,-1),14)]))
    story.append(hdr)

    sub = Table([[ps('T.P. 124567-T · info@contablearistizabal.com.co · Tel: 315-890-4321', 8, color=WHITE),
                  ps(f'Generado: {date.today().strftime("%d/%m/%Y")}', 8, color=WHITE, align=2)]],
                colWidths=['*', '*'])
    sub.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),BLUE),
        ('TOPPADDING',(0,0),(-1,-1),5),('BOTTOMPADDING',(0,0),(-1,-1),5),
        ('LEFTPADDING',(0,0),(-1,-1),14),('RIGHTPADDING',(0,0),(-1,-1),14)]))
    story.append(sub)
    story.append(Spacer(1, 0.35*cm))

    # Empresa info
    ei = Table([[ps(e['razon_social'], 12, True), ps(f"NIT: {e['nit']}", 10, align=2)],
                [ps(f"{e['sector']} · {e['ciudad']}", 8, color=MGRAY),
                 ps('Período: Ene – Jun 2026', 8, color=MGRAY, align=2)]],
               colWidths=['*', '*'])
    ei.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,-1),LGRAY),
        ('TOPPADDING',(0,0),(-1,0),10),('BOTTOMPADDING',(0,-1),(-1,-1),10),
        ('TOPPADDING',(0,1),(-1,1),3),('BOTTOMPADDING',(0,0),(-1,0),3),
        ('LEFTPADDING',(0,0),(-1,-1),14),('RIGHTPADDING',(0,0),(-1,-1),14),
        ('LINEBELOW',(0,-1),(-1,-1),0.8,BLUE)]))
    story.append(ei)
    story.append(Spacer(1, 0.45*cm))

    # KPIs
    utilidad = vt['neto'] - gt['neto']
    kpi = Table([
        [ps('VENTAS NETAS', 8, color=MGRAY, align=1), ps('POR COBRAR', 8, color=MGRAY, align=1),
         ps('CARTERA VENCIDA', 8, color=MGRAY, align=1), ps('UTILIDAD BRUTA', 8, color=MGRAY, align=1)],
        [ps(fmt(vt['neto']), 12, True, BLUE, 1), ps(fmt(vt['por_cobrar']), 12, True, DARK, 1),
         ps(fmt(vt['vencido']), 12, True, RED, 1), ps(fmt(utilidad), 12, True, GREEN if utilidad > 0 else RED, 1)],
    ], colWidths=['*', '*', '*', '*'])
    kpi.setStyle(TableStyle([
        ('BOX',(0,0),(0,-1),0.8,BLUE),('BOX',(1,0),(1,-1),0.5,colors.HexColor('#cbd5e1')),
        ('BOX',(2,0),(2,-1),0.8,RED),('BOX',(3,0),(3,-1),0.8,GREEN),
        ('TOPPADDING',(0,0),(-1,-1),8),('BOTTOMPADDING',(0,0),(-1,-1),8),
        ('LEFTPADDING',(0,0),(-1,-1),4),('RIGHTPADDING',(0,0),(-1,-1),4)]))
    story.append(kpi)
    story.append(Spacer(1, 0.55*cm))

    # Ventas table
    story.append(ps('FACTURAS DE VENTA', 10, True, NAVY))
    story.append(Spacer(1, 0.18*cm))
    vh = [ps(h, 8, True, WHITE, 1) for h in ['N°','Fecha','Cliente','Subtotal','IVA','Retefuente','Total','Neto','Estado']]
    vrows = [vh]
    for r in vs:
        ec = RED if 'VENCIDA' in str(r['estado']) else (colors.HexColor('#d97706') if r['estado']=='POR_VENCER' else (GREEN if r['estado']=='PAGADA' else DARK))
        vrows.append([ps(r['numero'],7,align=1), ps(r['fecha'],7,align=1),
            ps((r['cliente_nombre'] or '')[:24],7), ps(fmt(r['subtotal']),7,align=2),
            ps(fmt(r['iva']),7,align=2), ps(fmt(r['retefuente']),7,align=2),
            ps(fmt(r['total_factura']),7,align=2), ps(fmt(r['valor_neto']),7,True,align=2),
            ps(r['estado'][:14],7,color=ec,align=1)])
    vrows.append([ps('TOTAL',8,True,align=1), ps('',7), ps('',7),
        ps(fmt(sum(r['subtotal'] or 0 for r in vs)),8,True,align=2),
        ps(fmt(sum(r['iva'] or 0 for r in vs)),8,True,align=2),
        ps(fmt(vt['rf']),8,True,RED,2), ps(fmt(vt['total']),8,True,align=2),
        ps(fmt(vt['neto']),8,True,BLUE,2), ps('',7)])
    vtbl = Table(vrows, colWidths=[1.6*cm,1.7*cm,3.9*cm,1.9*cm,1.5*cm,1.9*cm,1.9*cm,1.9*cm,1.7*cm])
    vtbl.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),NAVY),('BACKGROUND',(0,-1),(-1,-1),LGRAY),
        ('ROWBACKGROUNDS',(0,1),(-1,-2),[WHITE,colors.HexColor('#f8fafc')]),
        ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#e2e8f0')),
        ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
        ('LEFTPADDING',(0,0),(-1,-1),3),('RIGHTPADDING',(0,0),(-1,-1),3),
        ('LINEBELOW',(0,-1),(-1,-1),1.2,BLUE)]))
    story.append(vtbl)
    story.append(Spacer(1, 0.55*cm))

    # Gastos table
    story.append(ps('FACTURAS DE GASTOS', 10, True, NAVY))
    story.append(Spacer(1, 0.18*cm))
    gh = [ps(h, 8, True, WHITE, 1) for h in ['N°','Fecha','Proveedor','Categoría','Subtotal','IVA','Retefuente','Total','Neto']]
    grows = [gh]
    for r in gs:
        grows.append([ps(r['numero'],7,align=1), ps(r['fecha'],7,align=1),
            ps((r['proveedor_nombre'] or '')[:20],7), ps((r['categoria'] or '')[:12],7),
            ps(fmt(r['subtotal']),7,align=2), ps(fmt(r['iva']),7,align=2),
            ps(fmt(r['retefuente']),7,align=2), ps(fmt(r['total_factura']),7,align=2),
            ps(fmt(r['valor_neto']),7,True,align=2)])
    grows.append([ps('TOTAL',8,True,align=1), ps('',7), ps('',7), ps('',7),
        ps(fmt(sum(r['subtotal'] or 0 for r in gs)),8,True,align=2),
        ps(fmt(sum(r['iva'] or 0 for r in gs)),8,True,align=2),
        ps(fmt(gt['rf']),8,True,RED,2), ps(fmt(gt['total']),8,True,align=2),
        ps(fmt(gt['neto']),8,True,RED,2)])
    gtbl = Table(grows, colWidths=[1.6*cm,1.7*cm,3.7*cm,2.1*cm,1.9*cm,1.5*cm,1.9*cm,1.8*cm,1.8*cm])
    gtbl.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),NAVY),('BACKGROUND',(0,-1),(-1,-1),LGRAY),
        ('ROWBACKGROUNDS',(0,1),(-1,-2),[WHITE,colors.HexColor('#f8fafc')]),
        ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#e2e8f0')),
        ('TOPPADDING',(0,0),(-1,-1),3),('BOTTOMPADDING',(0,0),(-1,-1),3),
        ('LEFTPADDING',(0,0),(-1,-1),3),('RIGHTPADDING',(0,0),(-1,-1),3),
        ('LINEBELOW',(0,-1),(-1,-1),1.2,RED)]))
    story.append(gtbl)
    story.append(Spacer(1, 0.55*cm))

    # Retenciones
    story.append(ps('RETENCIONES DEL PERÍODO', 10, True, NAVY))
    story.append(Spacer(1, 0.18*cm))
    rrows = [
        [ps(h, 9, True, WHITE, align) for h, align in [('Concepto',0),('Retefuente',2),('ReteIVA',2),('ReteICA',2),('Total',2)]],
        [ps('Practicadas (retenidas a proveedores)',9),
         ps(fmt(gt['rf']),9,align=2), ps(fmt(gt['ri']),9,align=2), ps(fmt(gt['rc']),9,align=2),
         ps(fmt(gt['rf']+gt['ri']+gt['rc']),9,True,RED,2)],
        [ps('Sufridas (retenidas por clientes)',9),
         ps(fmt(vt['rf']),9,align=2), ps(fmt(vt['ri']),9,align=2), ps(fmt(vt['rc']),9,align=2),
         ps(fmt(vt['rf']+vt['ri']+vt['rc']),9,True,BLUE,2)],
    ]
    saldo = (gt['rf']+gt['ri']+gt['rc']) - (vt['rf']+vt['ri']+vt['rc'])
    rrows.append([ps('Saldo a declarar este período',9,True),
        ps('',9), ps('',9), ps('',9),
        ps(fmt(saldo),10,True,RED if saldo > 0 else GREEN, 2)])
    rtbl = Table(rrows, colWidths=[6.5*cm, 2.4*cm, 2.4*cm, 2.4*cm, 2.3*cm])
    rtbl.setStyle(TableStyle([('BACKGROUND',(0,0),(-1,0),NAVY),('BACKGROUND',(0,-1),(-1,-1),LGRAY),
        ('ROWBACKGROUNDS',(0,1),(-1,-2),[WHITE,colors.HexColor('#f8fafc')]),
        ('GRID',(0,0),(-1,-1),0.25,colors.HexColor('#e2e8f0')),
        ('TOPPADDING',(0,0),(-1,-1),6),('BOTTOMPADDING',(0,0),(-1,-1),6),
        ('LEFTPADDING',(0,0),(-1,-1),8),('RIGHTPADDING',(0,0),(-1,-1),8),
        ('LINEBELOW',(0,-1),(-1,-1),1.2,BLUE)]))
    story.append(rtbl)
    story.append(Spacer(1, 0.5*cm))

    # Footer
    story.append(HRFlowable(width='100%', thickness=0.4, color=MGRAY))
    story.append(Spacer(1, 0.15*cm))
    story.append(ps('Informe generado automáticamente por ContaBot. Debe ser verificado por el contador '
                    'responsable antes de ser presentado como soporte oficial ante la DIAN u otras entidades.', 7, color=MGRAY))

    doc.build(story)
    buf.seek(0)
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/pdf'
    nombre_safe = re.sub(r'[^a-zA-Z0-9]', '_', e['razon_social'])[:28]
    resp.headers['Content-Disposition'] = f'attachment; filename="Informe_{nombre_safe}_2026.pdf"'
    return resp


# ── Pre-liquidación de retenciones DIAN ──────────────────────────────────────

@app.route("/api/declaraciones")
@login_required
def declaraciones():
    from calendario import obligaciones_retefte
    cid = session["contador_id"]
    emps = (
        sb.table("empresas_clientes")
        .select("id,razon_social,nit,color,regimen")
        .eq("contador_id", cid)
        .order("id")
        .execute().data
    )
    hoy = date.today()
    MESES_ES = ["enero","febrero","marzo","abril","mayo","junio",
                "julio","agosto","septiembre","octubre","noviembre","diciembre"]
    mes_actual_label = f"{MESES_ES[hoy.month-1].capitalize()} {hoy.year}"

    resultado    = []
    total_global = 0

    # Batch: 2 queries para todas las empresas en lugar de 2N
    emp_ids = [e["id"] for e in emps]
    if emp_ids:
        g_all = sb.table("facturas_gastos").select("empresa_id,retefuente,reteiva,reteica").in_("empresa_id", emp_ids).execute().data
        v_all = sb.table("facturas_venta").select("empresa_id,retefuente,reteiva,reteica").in_("empresa_id", emp_ids).execute().data
    else:
        g_all, v_all = [], []
    g_by_emp = {}
    v_by_emp = {}
    for r in g_all: g_by_emp.setdefault(r["empresa_id"], []).append(r)
    for r in v_all: v_by_emp.setdefault(r["empresa_id"], []).append(r)

    for e in emps:
        regimen = e.get("regimen") or "Juridica"
        aplica_rtefte = regimen in ("Juridica", "GranContribuyente")

        g_rows = g_by_emp.get(e["id"], [])
        v_rows = v_by_emp.get(e["id"], [])

        rf = round(sum(r["retefuente"] or 0 for r in g_rows)) if aplica_rtefte else 0
        ri = round(sum(r["reteiva"]    or 0 for r in g_rows)) if aplica_rtefte else 0
        rc = round(sum(r["reteica"]    or 0 for r in g_rows)) if aplica_rtefte else 0
        total = rf + ri + rc
        total_global += total

        # Fecha vencimiento real según NIT y calendario DIAN
        if aplica_rtefte:
            obs_rtefte = obligaciones_retefte(e["nit"])
            # Buscar el vencimiento del mes actual o siguiente
            fecha_limite_str = next(
                (o["vencimiento"] for o in obs_rtefte if o["vencimiento"] >= hoy.isoformat()),
                obs_rtefte[-1]["vencimiento"] if obs_rtefte else hoy.isoformat()
            )
            fecha_limite = date.fromisoformat(fecha_limite_str)
            dias = (fecha_limite - hoy).days
            # Formatear fecha en español
            fl_label = f"{fecha_limite.day} de {MESES_ES[fecha_limite.month-1]} {fecha_limite.year}"
        else:
            fecha_limite_str = None
            dias = None
            fl_label = None

        resultado.append({
            "empresa_id":         e["id"],
            "razon_social":       e["razon_social"],
            "nit":                e["nit"],
            "color":              e["color"] or "#6366f1",
            "regimen":            regimen,
            "aplica_rtefte":      aplica_rtefte,
            "retefuente":         rf,
            "reteiva":            ri,
            "reteica":            rc,
            "total":              total,
            "sufrido_retefuente": round(sum(r["retefuente"] or 0 for r in v_rows)),
            "fecha_limite":       fecha_limite_str,
            "fecha_limite_label": fl_label,
            "dias":               dias,
            "estado":             ("VENCIDA" if dias is not None and dias < 0
                                   else "HOY" if dias == 0
                                   else "PENDIENTE" if dias is not None
                                   else "NO_APLICA"),
        })

    return jsonify({
        "mes":               mes_actual_label,
        "empresas":          resultado,
        "total_consolidado": total_global,
    })


# ── Flujo de caja proyectado 60 días ─────────────────────────────────────────

@app.route("/api/empresa/<int:eid>/flujo-caja")
@login_required
def flujo_caja(eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    hoy = date.today()

    ventas_rows = (
        sb.table("facturas_venta")
        .select("fecha_vencimiento,valor_neto")
        .eq("empresa_id", eid)
        .in_("estado", ["PENDIENTE", "POR_VENCER"])
        .gte("fecha_vencimiento", hoy.isoformat())
        .execute().data
    )
    gastos_rows = (
        sb.table("facturas_gastos")
        .select("fecha_vencimiento,valor_neto")
        .eq("empresa_id", eid)
        .in_("estado", ["PENDIENTE", "POR_VENCER"])
        .gte("fecha_vencimiento", hoy.isoformat())
        .execute().data
    )
    cartera_rows = (
        sb.table("facturas_venta")
        .select("valor_neto")
        .eq("empresa_id", eid)
        .like("estado", "VENCIDA%")
        .execute().data
    )
    cartera_total = sum(r["valor_neto"] or 0 for r in cartera_rows)

    semanas = []
    for i in range(8):
        ini = hoy + timedelta(days=i * 7)
        fin = ini + timedelta(days=6)
        ing = sum(r["valor_neto"] or 0 for r in ventas_rows
                  if ini <= date.fromisoformat(r["fecha_vencimiento"]) <= fin)
        egr = sum(r["valor_neto"] or 0 for r in gastos_rows
                  if ini <= date.fromisoformat(r["fecha_vencimiento"]) <= fin)
        semanas.append({
            "label":    f"{ini.strftime('%d/%m')}–{fin.strftime('%d/%m')}",
            "ingresos": round(ing),
            "egresos":  round(egr),
            "neto":     round(ing - egr),
        })

    return jsonify({"semanas": semanas, "cartera_vencida": round(cartera_total)})


@app.route("/api/empresa/<int:eid>/declaraciones")
@login_required
def declaraciones_empresa(eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    from datetime import date as d
    e_rows = sb.table("empresas_clientes").select("*").eq("id", eid).execute().data
    if not e_rows:
        return jsonify({"error": "Empresa no encontrada"}), 404
    e = e_rows[0]

    ventas = sb.table("facturas_venta").select("fecha,subtotal,iva,retefuente,reteiva,reteica,total_factura,valor_neto,tipo_documento").eq("empresa_id", eid).execute().data
    gastos = sb.table("facturas_gastos").select("fecha,subtotal,iva,retefuente,reteiva,reteica,total_factura,valor_neto").eq("empresa_id", eid).execute().data

    def mes(r):
        try: return int((r.get("fecha") or "")[:7].split("-")[1])
        except: return 0

    def cuatrimestre(r):
        m = mes(r)
        if 1 <= m <= 4: return 1
        if 5 <= m <= 8: return 2
        return 3

    def bimestre(r):
        m = mes(r)
        return (m + 1) // 2 if m > 0 else 0

    # NC resta de la base (signo -1), factura normal suma (signo +1)
    def signo(r): return -1 if r.get("tipo_documento") == "nota_credito" else 1

    # IVA F-300 por cuatrimestre
    f300 = []
    labels_c = {1: "Ene–Abr", 2: "May–Ago", 3: "Sep–Dic"}
    for c in [1, 2, 3]:
        vv = [r for r in ventas if cuatrimestre(r) == c]
        gg = [r for r in gastos if cuatrimestre(r) == c]
        iva_gen  = round(sum(signo(r) * (r["iva"] or 0) for r in vv))
        iva_desc = round(sum(r["iva"] or 0 for r in gg))
        iva_pagar = max(0, iva_gen - iva_desc)
        f300.append({
            "cuatrimestre": c,
            "periodo":      labels_c[c],
            "base_ventas":  round(sum(signo(r) * (r["subtotal"] or 0) for r in vv)),
            "iva_generado": iva_gen,
            "iva_descontable": iva_desc,
            "iva_a_pagar":  iva_pagar,
            "n_facturas_v": len(vv),
            "n_facturas_g": len(gg),
        })

    # Retefuente F-350 por mes
    MESES = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    f350 = []
    for m_num in range(1, 13):
        gg = [r for r in gastos if mes(r) == m_num]
        rfte = round(sum(r["retefuente"] or 0 for r in gg))
        if rfte > 0 or gg:
            f350.append({
                "mes":       m_num,
                "periodo":   MESES[m_num - 1],
                "base":      round(sum(r["subtotal"] or 0 for r in gg)),
                "retefte":   rfte,
                "n_facturas": len(gg),
            })

    # ICA por bimestre
    TASA_ICA = 4.14 / 1000  # 4.14‰ Bogotá (ajustable por ciudad)
    labels_b = {1:"Ene–Feb",2:"Mar–Abr",3:"May–Jun",4:"Jul–Ago",5:"Sep–Oct",6:"Nov–Dic"}
    ica = []
    for b in range(1, 7):
        vv = [r for r in ventas if bimestre(r) == b]
        base = round(sum(signo(r) * (r["subtotal"] or 0) for r in vv))
        ica.append({
            "bimestre":   b,
            "periodo":    labels_b[b],
            "base":       base,
            "tasa":       "4.14‰",
            "ica_a_pagar": round(base * TASA_ICA),
            "n_facturas": len(vv),
        })

    regimen = e.get("regimen") or "Juridica"
    # responsable_iva = jurídica obligada a declarar IVA (pero no necesariamente agente retenedor)
    aplica_iva    = regimen in ("Juridica", "GranContribuyente", "responsable_iva")
    aplica_rtefte = regimen in ("Juridica", "GranContribuyente")
    aplica_ica    = regimen in ("Juridica", "GranContribuyente", "Simple", "responsable_iva")

    # Renta anual — estimación a partir de facturas electrónicas
    ingresos_renta      = round(sum(signo(r) * (r["valor_neto"] or 0) for r in ventas))
    gastos_deducibles   = round(sum(r["valor_neto"] or 0 for r in gastos))
    renta_liquida       = ingresos_renta - gastos_deducibles
    es_juridica_renta   = regimen in ("Juridica", "GranContribuyente")
    tasa_impuesto       = 35 if es_juridica_renta else 0
    impuesto_estimado   = round(renta_liquida * 0.35) if es_juridica_renta and renta_liquida > 0 else 0

    if regimen == "Natural":
        checklist_renta = [
            "Patrimonio bruto (inmuebles, vehículos, cuentas bancarias, inversiones a dic 31)",
            "Pasivos y deudas vigentes",
            "Ingresos no facturados (arriendos, intereses, dividendos, salarios)",
            "Gastos deducibles no facturados (salud, educación, dependientes)",
            "Certificados de retención en la fuente recibidos",
            "Extractos bancarios del período",
        ]
    else:
        checklist_renta = [
            "Estado de resultados completo del período",
            "Balance general (activos y pasivos)",
            "Provisiones contables",
            "Deducciones especiales (donaciones, inversiones)",
            "Retenciones en la fuente que le practicaron",
        ]

    renta = {
        "ingresos":               ingresos_renta,
        "gastos_deducibles":      gastos_deducibles,
        "renta_liquida_estimada": renta_liquida,
        "tasa_impuesto":          tasa_impuesto,
        "impuesto_estimado":      impuesto_estimado,
        "checklist":              checklist_renta,
    }

    return jsonify({
        "ok":      True,
        "regimen": regimen,
        "empresa": {"id": e["id"], "razon_social": e["razon_social"], "nit": e["nit"]},
        "f300":  f300 if aplica_iva else [],
        "f350":  f350 if aplica_rtefte else [],
        "ica":   ica  if aplica_ica else [],
        "renta": renta,
        "aplica_iva":    aplica_iva,
        "aplica_rtefte": aplica_rtefte,
        "aplica_ica":    aplica_ica,
    })


# ── Export Excel — Formato Eduardo (CONTABILIDAD 2026) ────────────────────────

def _mes_num(fecha):
    try: return int(str(fecha or "")[:7].split("-")[1])
    except: return 0

def _bimestre(m): return (m + 1) // 2 if m else 0

def _cuatrimestre(m):
    if 1 <= m <= 4: return 1
    if 5 <= m <= 8: return 2
    return 3 if m else 0

def _cod_iva(subtotal, iva):
    if not subtotal or subtotal == 0: return "EX0 EXENTO"
    tasa = (iva or 0) / subtotal
    if abs(tasa - 0.19) < 0.02: return "IVA 19 GEN"
    if abs(tasa - 0.05) < 0.02: return "IVA 5 GEN"
    return "EX0 EXENTO"

def _pct_iva(subtotal, iva):
    if not subtotal or subtotal == 0: return 0
    return round((iva or 0) / subtotal, 4)


@app.route("/api/empresa/<int:eid>/informe-excel")
@login_required
@limiter.limit("10 per minute")
def informe_excel(eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        return jsonify({"error": "pip install openpyxl"}), 500
    import io
    from flask import make_response
    from datetime import date as d

    e_rows = sb.table("empresas_clientes").select("*").eq("id", eid).execute().data
    if not e_rows:
        return jsonify({"error": "Empresa no encontrada"}), 404
    e = e_rows[0]

    vs = sb.table("facturas_venta").select(
        "numero,cufe,fecha,cliente_nombre,cliente_nit,subtotal,iva,"
        "retefuente,reteiva,reteica,total_factura,valor_neto,estado,tipo_documento"
    ).eq("empresa_id", eid).order("fecha").execute().data

    gs = sb.table("facturas_gastos").select(
        "numero,cufe,fecha,proveedor_nombre,proveedor_nit,categoria,"
        "subtotal,iva,retefuente,reteiva,reteica,total_factura,valor_neto,estado"
    ).eq("empresa_id", eid).order("fecha").execute().data

    # ── helpers de estilo ──────────────────────────────────────────────────────
    HDR_FILL = PatternFill("solid", fgColor="1E3A5F")
    HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    NUM_FMT = '#,##0'

    def style_header(ws, cols, row=1):
        for ci, col in enumerate(cols, 1):
            cell = ws.cell(row=row, column=ci, value=col)
            cell.font = HDR_FONT; cell.fill = HDR_FILL
            cell.alignment = HDR_ALIGN; cell.border = BORDER

    def style_row(ws, row_num, n_cols, is_num_cols=None):
        fill = PatternFill("solid", fgColor="F8FAFC") if row_num % 2 == 0 else None
        for ci in range(1, n_cols + 1):
            cell = ws.cell(row=row_num, column=ci)
            cell.border = BORDER
            if fill: cell.fill = fill
            if is_num_cols and ci in is_num_cols:
                cell.number_format = NUM_FMT
                cell.alignment = Alignment(horizontal="right")

    wb = openpyxl.Workbook()

    # ── Hoja VENTAS ────────────────────────────────────────────────────────────
    ws_v = wb.active; ws_v.title = "VENTAS"
    ws_v.row_dimensions[1].height = 40

    title = f"REGISTRO DE VENTAS — {e['razon_social']} | NIT {e['nit']} | 2026"
    ws_v.merge_cells("A1:W1")
    tc = ws_v["A1"]; tc.value = title
    tc.font = Font(bold=True, size=12, color="1E3A5F")
    tc.alignment = Alignment(horizontal="center", vertical="center")

    V_COLS = ["Nº","Fecha","Mes","Bimestre","Cuatrimestre","Nº Factura","CUFE",
              "Cliente","NIT","Concepto","Contrato/Orden","Cuenta PUC",
              "Base Gravable","Cód. IVA","% IVA","IVA Generado",
              "RteFte 135515","ReteIVA 135517","ReteICA 135518",
              "Total Factura","Neto a Recibir","Estado Cobro","Observaciones"]
    style_header(ws_v, V_COLS, row=2)

    NUM_COLS_V = {13,16,17,18,19,20,21}
    for i, r in enumerate(vs, 1):
        m = _mes_num(r["fecha"])
        row_data = [
            i, r["fecha"], m, _bimestre(m), _cuatrimestre(m),
            r["numero"], r.get("cufe",""),
            r["cliente_nombre"], r.get("cliente_nit",""),
            r.get("concepto",""), "", "",
            r["subtotal"] or 0,
            _cod_iva(r["subtotal"], r["iva"]),
            _pct_iva(r["subtotal"], r["iva"]),
            r["iva"] or 0,
            r["retefuente"] or 0, r["reteiva"] or 0, r["reteica"] or 0,
            r["total_factura"] or 0, r["valor_neto"] or 0,
            r["estado"], ""
        ]
        for ci, val in enumerate(row_data, 1):
            ws_v.cell(row=i+2, column=ci, value=val)
        style_row(ws_v, i+2, len(V_COLS), NUM_COLS_V)

    # Anchos columna VENTAS
    for ci, w in enumerate([5,12,5,8,10,14,20,28,14,20,16,12,
                             14,12,8,14,14,14,14,14,14,12,16], 1):
        ws_v.column_dimensions[get_column_letter(ci)].width = w

    # ── Hoja COMPRAS ───────────────────────────────────────────────────────────
    ws_g = wb.create_sheet("COMPRAS")
    ws_g.row_dimensions[1].height = 40
    ws_g.merge_cells("A1:Z1")
    tc2 = ws_g["A1"]; tc2.value = f"REGISTRO DE COMPRAS — {e['razon_social']} | NIT {e['nit']} | 2026"
    tc2.font = Font(bold=True, size=12, color="1E3A5F")
    tc2.alignment = Alignment(horizontal="center", vertical="center")

    G_COLS = ["Nº","Fecha","Mes","Bimestre","Cuatrimestre","Nº Factura",
              "Proveedor","NIT","Concepto / Detalle","Tipo (Bien/Servicio)","CC","Cuenta PUC",
              "Base Gravable","Cód. IVA","% IVA","VALOR IVA",
              "Cód. RteFte","% RteFte","Valor RteFte","% ReteICA","Valor ReteICA",
              "Total Factura","Neto a Pagar","Forma Pago","Estado","Observaciones"]
    style_header(ws_g, G_COLS, row=2)

    NUM_COLS_G = {13,16,19,21,22,23}
    TIPO_MAP = {"honorarios":"Servicio","servicios":"Servicio","transporte":"Servicio",
                "insumos":"Bien","tecnologia":"Bien"}
    for i, r in enumerate(gs, 1):
        m = _mes_num(r["fecha"])
        cat = (r.get("categoria") or "").lower()
        tipo = TIPO_MAP.get(cat, "Servicio")
        pct_rf = round((r["retefuente"] or 0) / (r["subtotal"] or 1), 4) if r.get("subtotal") else 0
        row_data = [
            i, r["fecha"], m, _bimestre(m), _cuatrimestre(m),
            r["numero"], r["proveedor_nombre"], r.get("proveedor_nit",""),
            r.get("categoria",""), tipo, "", "",
            r["subtotal"] or 0,
            _cod_iva(r["subtotal"], r["iva"]),
            _pct_iva(r["subtotal"], r["iva"]),
            r["iva"] or 0,
            "", pct_rf, r["retefuente"] or 0,
            0, r["reteica"] or 0,
            r["total_factura"] or 0, r["valor_neto"] or 0,
            "", r["estado"], ""
        ]
        for ci, val in enumerate(row_data, 1):
            ws_g.cell(row=i+2, column=ci, value=val)
        style_row(ws_g, i+2, len(G_COLS), NUM_COLS_G)

    for ci, w in enumerate([5,12,5,8,10,14,28,14,22,12,8,12,
                             14,12,8,14,14,8,14,8,12,14,14,12,12,16], 1):
        ws_g.column_dimensions[get_column_letter(ci)].width = w

    # ── Hoja IVA F-300 ─────────────────────────────────────────────────────────
    ws_iva = wb.create_sheet("IVA F-300")
    ws_iva.merge_cells("A1:F1")
    h = ws_iva["A1"]; h.value = f"DECLARACIÓN IVA F-300 — {e['razon_social']} | 2026"
    h.font = Font(bold=True, size=12, color="1E3A5F")
    h.alignment = Alignment(horizontal="center")
    style_header(ws_iva, ["Concepto","Cuatrimestre 1\n(Ene-Abr)",
                           "Cuatrimestre 2\n(May-Ago)","Cuatrimestre 3\n(Sep-Dic)",
                           "TOTAL AÑO","Casilla F-300"], row=2)

    def cuatri_data(rows, q):
        return [r for r in rows if _cuatrimestre(_mes_num(r["fecha"])) == q]

    def _sgn(r): return -1 if r.get("tipo_documento") == "nota_credito" else 1

    conceptos_iva = [
        ("Base gravable ventas",         lambda q: sum(_sgn(r) * (r["subtotal"] or 0) for r in cuatri_data(vs,q))),
        ("IVA generado (ventas)",         lambda q: sum(_sgn(r) * (r["iva"] or 0) for r in cuatri_data(vs,q))),
        ("Base gravable compras",         lambda q: sum(r["subtotal"] or 0 for r in cuatri_data(gs,q))),
        ("IVA descontable (compras)",     lambda q: sum(r["iva"] or 0 for r in cuatri_data(gs,q))),
        ("IVA a pagar (Generado-Desct.)", lambda q: max(0, sum(_sgn(r) * (r["iva"] or 0) for r in cuatri_data(vs,q)) - sum(r["iva"] or 0 for r in cuatri_data(gs,q)))),
    ]
    CASILLAS = ["","","","","67"]
    for ri, (concepto, fn) in enumerate(conceptos_iva, 3):
        vals = [fn(1), fn(2), fn(3)]
        row_data = [concepto, *vals, sum(vals), CASILLAS[ri-3]]
        for ci, val in enumerate(row_data, 1):
            cell = ws_iva.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            if ci > 1 and ci < 6 and isinstance(val, (int,float)):
                cell.number_format = NUM_FMT
        if concepto.startswith("IVA a pagar"):
            for ci in range(2, 6):
                ws_iva.cell(row=ri, column=ci).font = Font(bold=True, color="C0392B")

    # ── Hoja RFTE 350 ──────────────────────────────────────────────────────────
    ws_rf = wb.create_sheet("RFTE 350")
    ws_rf.merge_cells("A1:O1")
    h2 = ws_rf["A1"]; h2.value = f"RETENCIÓN EN LA FUENTE F-350 — {e['razon_social']} | 2026"
    h2.font = Font(bold=True, size=12, color="1E3A5F")
    h2.alignment = Alignment(horizontal="center")
    MESES_NOM = ["Ene","Feb","Mar","Abr","May","Jun","Jul","Ago","Sep","Oct","Nov","Dic"]
    style_header(ws_rf, ["Concepto","Código",*MESES_NOM,"Total Año"], row=2)

    def mes_data(rows, m): return [r for r in rows if _mes_num(r["fecha"]) == m]
    rf_vals = [round(sum(r["retefuente"] or 0 for r in mes_data(gs, m))) for m in range(1,13)]
    rf_row = ["Retención en la fuente practicada", "75-96", *rf_vals, sum(rf_vals)]
    for ci, val in enumerate(rf_row, 1):
        cell = ws_rf.cell(row=3, column=ci, value=val)
        cell.border = BORDER
        if ci > 2 and isinstance(val, (int,float)):
            cell.number_format = NUM_FMT
            if val > 0: cell.font = Font(bold=True, color="C0392B")

    # ── Hoja ICA ───────────────────────────────────────────────────────────────
    ws_ica = wb.create_sheet("ICA")
    ws_ica.merge_cells("A1:I1")
    h3 = ws_ica["A1"]; h3.value = f"DECLARACIÓN ICA — {e['razon_social']} | 2026"
    h3.font = Font(bold=True, size=12, color="1E3A5F")
    h3.alignment = Alignment(horizontal="center")
    BIM_LABELS = ["Bim 1\n(Ene-Feb)","Bim 2\n(Mar-Abr)","Bim 3\n(May-Jun)",
                  "Bim 4\n(Jul-Ago)","Bim 5\n(Sep-Oct)","Bim 6\n(Nov-Dic)"]
    style_header(ws_ica, ["Concepto",*BIM_LABELS,"TOTAL AÑO"], row=2)

    def bim_data(rows, b): return [r for r in rows if _bimestre(_mes_num(r["fecha"])) == b]
    TASA_ICA = 4.14 / 1000
    ica_bases = [round(sum(_sgn(r) * (r["subtotal"] or 0) for r in bim_data(vs, b))) for b in range(1,7)]
    ica_vals  = [round(b * TASA_ICA) for b in ica_bases]

    for ri, (label, vals) in enumerate([("Base gravable (ventas)", ica_bases),
                                         ("ICA a pagar (4.14‰)", ica_vals)], 3):
        row_data = [label, *vals, sum(vals)]
        for ci, val in enumerate(row_data, 1):
            cell = ws_ica.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            if ci > 1 and isinstance(val,(int,float)):
                cell.number_format = NUM_FMT
        if label.startswith("ICA a pagar"):
            for ci in range(2, 9):
                ws_ica.cell(row=ri, column=ci).font = Font(bold=True, color="C0392B")

    # ── Hoja RENTA ─────────────────────────────────────────────────────────────
    regimen_excel = e.get("regimen") or "Juridica"
    ws_r = wb.create_sheet("RENTA")
    ws_r.merge_cells("A1:C1")
    hr = ws_r["A1"]; hr.value = f"ESTIMACIÓN RENTA ANUAL — {e['razon_social']} | 2026"
    hr.font = Font(bold=True, size=12, color="1E3A5F")
    hr.alignment = Alignment(horizontal="center")
    style_header(ws_r, ["Concepto", "Valor (COP)", "Nota"], row=2)
    es_jur = regimen_excel in ("Juridica", "GranContribuyente", "responsable_iva")
    ingr   = round(sum(_sgn(r) * (r["valor_neto"] or 0) for r in vs))
    gast   = round(sum(r["valor_neto"] or 0 for r in gs))
    base   = ingr - gast
    imp    = round(base * 0.35) if es_jur and base > 0 else 0
    renta_rows = [
        ("Ingresos identificados (facturas electrónicas)", ingr, "Ventas netas del período"),
        ("Gastos deducibles identificados (facturas)",     gast, "Compras y gastos del período"),
        ("Base estimada",                                  base, "Solo facturas en ContaBot"),
    ]
    if es_jur:
        renta_rows.append(("Impuesto estimado (35%)", imp, "Tarifa general personas jurídicas"))
    for ri, (concepto, valor, nota) in enumerate(renta_rows, 3):
        ws_r.cell(row=ri, column=1, value=concepto).border = BORDER
        vc = ws_r.cell(row=ri, column=2, value=valor)
        vc.border = BORDER; vc.number_format = NUM_FMT; vc.alignment = Alignment(horizontal="right")
        if concepto.startswith("Base") or concepto.startswith("Impuesto"):
            vc.font = Font(bold=True, color="C0392B" if valor > 0 else "27AE60")
        ws_r.cell(row=ri, column=3, value=nota).border = BORDER
    # Fila de advertencia
    ri_adv = len(renta_rows) + 4
    ws_r.merge_cells(f"A{ri_adv}:C{ri_adv}")
    adv = ws_r.cell(row=ri_adv, column=1,
        value="⚠️ Esta estimación solo incluye facturas electrónicas registradas en ContaBot. "
              "Para declarar renta se requiere: patrimonio, otros ingresos, deducciones personales y certificados de retención.")
    adv.font = Font(italic=True, color="7F8C8D", size=9)
    adv.alignment = Alignment(wrap_text=True)
    ws_r.row_dimensions[ri_adv].height = 40
    ws_r.column_dimensions["A"].width = 48
    ws_r.column_dimensions["B"].width = 18
    ws_r.column_dimensions["C"].width = 36

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = make_response(buf.read())
    resp.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    nombre = re.sub(r'[^a-zA-Z0-9]', '_', e['razon_social'])[:28]
    resp.headers['Content-Disposition'] = f'attachment; filename="ContaBot_{nombre}_2026.xlsx"'
    return resp


def _normalizar_fecha(val) -> str:
    """Convierte cualquier valor de fecha a YYYY-MM-DD para Supabase."""
    import re as _re2
    from datetime import datetime as _dt
    if val is None:
        return ""
    if hasattr(val, 'strftime'):
        return val.strftime("%Y-%m-%d")
    s = str(val).strip()[:10]
    # DD-MM-YYYY o DD/MM/YYYY → YYYY-MM-DD
    m = _re2.match(r'^(\d{2})[-/](\d{2})[-/](\d{4})$', s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return s


# ── Importar Excel DIAN — conciliación ────────────────────────────────────────

@app.route("/api/empresa/<int:eid>/importar-dian", methods=["POST"])
@login_required
def importar_dian(eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    import io, re as _re
    try:
        import openpyxl
    except ImportError:
        return jsonify({"ok": False, "error": "pip install openpyxl"}), 500

    archivo = request.files.get("archivo")
    if not archivo:
        return jsonify({"ok": False, "error": "No se recibió archivo"}), 400

    try:
        wb = openpyxl.load_workbook(io.BytesIO(archivo.read()), read_only=True, data_only=True)
    except Exception as ex:
        return jsonify({"ok": False, "error": f"No se pudo leer el Excel: {ex}"}), 400

    CUFE_RE = _re.compile(r"^[a-f0-9]{96}$", _re.IGNORECASE)

    # Buscar CUFEs en todas las hojas
    cufes_dian = {}   # cufe → {numero, fecha, nit_emisor, nombre_emisor, total}
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = None
        cufe_col = num_col = fecha_col = nit_col = nombre_col = total_col = None

        for ri, row in enumerate(ws.iter_rows(values_only=True)):
            if headers is None:
                # Detectar fila de encabezados
                row_str = [str(c or "").lower() for c in row]
                if any("cufe" in c for c in row_str):
                    headers = row_str
                    for ci, h in enumerate(headers):
                        if "cufe" in h: cufe_col = ci
                        elif any(x in h for x in ["factura","numero","número","n°"]): num_col = ci
                        elif "fecha" in h: fecha_col = ci
                        elif "nit" in h: nit_col = ci
                        elif any(x in h for x in ["razon","razón","nombre","emisor"]): nombre_col = ci
                        elif any(x in h for x in ["total","valor"]): total_col = ci
                    continue
                # Sin encabezados: buscar CUFEs directamente en celdas
                for ci, cell in enumerate(row):
                    val = str(cell or "").strip()
                    if CUFE_RE.match(val):
                        cufes_dian[val.lower()] = {"cufe": val.lower(), "numero": "", "fecha": "", "nit_emisor": "", "nombre_emisor": "", "total": 0}
                continue

            # Con encabezados detectados
            if cufe_col is not None and len(row) > cufe_col:
                val = str(row[cufe_col] or "").strip()
                if CUFE_RE.match(val):
                    cufes_dian[val.lower()] = {
                        "cufe":         val.lower(),
                        "numero":       str(row[num_col] or "") if num_col is not None else "",
                        "fecha":        _normalizar_fecha(row[fecha_col]) if fecha_col is not None else "",
                        "nit_emisor":   str(row[nit_col] or "") if nit_col is not None else "",
                        "nombre_emisor":str(row[nombre_col] or "") if nombre_col is not None else "",
                        "total":        (lambda v: float(v) if isinstance(v,(int,float)) else 0.0)(row[total_col]) if total_col is not None else 0,
                    }

    if not cufes_dian:
        return jsonify({"ok": False, "error": "No se encontraron CUFEs en el archivo. ¿Es el Excel exportado del portal DIAN?"}), 400

    # Comparar contra lo que ya está en ContaBot
    registradas = sb.table("facturas_gastos").select("cufe,numero").eq("empresa_id", eid).execute().data
    cufes_contabot = {(r.get("cufe") or "").lower() for r in registradas}

    ya_registradas = []
    nuevas         = []
    for cufe, info in cufes_dian.items():
        if cufe in cufes_contabot:
            ya_registradas.append(info)
        else:
            nuevas.append(info)

    return jsonify({
        "ok":            True,
        "total_dian":    len(cufes_dian),
        "ya_en_contabot": len(ya_registradas),
        "nuevas":        len(nuevas),
        "detalle_nuevas": nuevas[:50],
    })


@app.route("/api/empresa/<int:eid>/importar-dian/registrar", methods=["POST"])
@login_required
def importar_dian_registrar(eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    """Registra en ContaBot las facturas DIAN que aún no están."""
    from datetime import datetime as dt
    facturas = request.get_json() or []
    if not facturas:
        return jsonify({"ok": False, "error": "No hay facturas"}), 400

    registradas_antes = {
        (r.get("cufe") or "").lower()
        for r in sb.table("facturas_gastos").select("cufe").eq("empresa_id", eid).execute().data
    }

    def _safe_float(v):
        try:
            if v is None: return 0.0
            if isinstance(v, (int, float)): return float(v)
            return float(str(v).replace(".", "").replace(",", ".").replace("$", "").strip())
        except: return 0.0

    insertadas = 0
    errores    = 0
    for f in facturas:
        try:
            cufe = (f.get("cufe") or "").lower().strip()
            if not cufe or cufe in registradas_antes:
                continue
            numero = (f.get("numero") or "").strip() or cufe[:24]
            fecha  = _normalizar_fecha(f.get("fecha")) or dt.today().strftime("%Y-%m-%d")
            total  = _safe_float(f.get("total"))
            _insertar_factura_gasto(eid, {
                "numero":           numero,
                "cufe":             cufe,
                "fecha":            fecha,
                "proveedor_nit":    (f.get("nit_emisor") or "").strip(),
                "proveedor_nombre": (f.get("nombre_emisor") or "").strip(),
                "proveedor_ciudad": "",
                "subtotal":         total,
                "iva":              0,
                "total_factura":    total,
                "valor_neto":       total,
            }, fuente="dian")
            registradas_antes.add(cufe)
            insertadas += 1
        except Exception:
            logging.exception("Error insertando factura DIAN cufe=%s", f.get("cufe", "?"))
            errores += 1

    return jsonify({"ok": True, "insertadas": insertadas, "errores": errores})


# ── Marcar factura como pagada ─────────────────────────────────────────────────

@app.route("/api/factura/<tipo>/<path:numero>/empresa/<int:eid>/pagar", methods=["POST"])
@login_required
def marcar_pagada(tipo, numero, eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    if tipo not in ("venta", "gasto"):
        return jsonify({"ok": False, "error": "Tipo inválido"}), 400
    tabla = "facturas_venta" if tipo == "venta" else "facturas_gastos"
    try:
        sb.table(tabla).update({"estado": "PAGADA"}).eq("numero", numero).eq("empresa_id", eid).execute()
        _cache_invalidar(session["contador_id"])
        _cache_invalidar(session["contador_id"])
        return jsonify({"ok": True})
    except Exception as ex:
        logging.exception('Error en endpoint')
        return jsonify({"ok": False, "error": "Error interno"}), 400


# ── Revertir estado PAGADA → PENDIENTE ────────────────────────────────────────

@app.route("/api/factura/<tipo>/<path:numero>/empresa/<int:eid>/despagar", methods=["POST"])
@login_required
def desmarcar_pagada(tipo, numero, eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    if tipo not in ("venta", "gasto"):
        return jsonify({"ok": False, "error": "Tipo inválido"}), 400
    tabla = "facturas_venta" if tipo == "venta" else "facturas_gastos"
    try:
        sb.table(tabla).update({"estado": "PENDIENTE"}).eq("numero", numero).eq("empresa_id", eid).execute()
        _cache_invalidar(session["contador_id"])
        return jsonify({"ok": True})
    except Exception as ex:
        logging.exception('Error en endpoint')
        return jsonify({"ok": False, "error": "Error interno"}), 400


# ── Borrar factura ────────────────────────────────────────────────────────────

@app.route("/api/factura/<tipo>/<path:numero>/empresa/<int:eid>", methods=["DELETE"])
@login_required
def borrar_factura(tipo, numero, eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    if tipo not in ("venta", "gasto"):
        return jsonify({"ok": False, "error": "Tipo inválido"}), 400
    tabla = "facturas_venta" if tipo == "venta" else "facturas_gastos"
    try:
        # Obtener archivo antes de borrar para eliminarlo de Storage
        row = sb.table(tabla).select("archivo_pdf").eq("numero", numero).eq("empresa_id", eid).execute().data
        sb.table(tabla).delete().eq("numero", numero).eq("empresa_id", eid).execute()
        # Borrar archivo de Storage si es URL de Supabase
        if row and row[0].get("archivo_pdf", "").startswith("http"):
            try:
                url = row[0]["archivo_pdf"]
                # Extraer path dentro del bucket: todo lo que va después de /object/public/facturas/
                import re as _re
                m = _re.search(r"/object/public/facturas/(.+)", url)
                if m:
                    sb.storage.from_("facturas").remove([m.group(1)])
            except Exception:
                pass
        _cache_invalidar(session["contador_id"])
        return jsonify({"ok": True})
    except Exception as ex:
        logging.exception('Error en endpoint')
        return jsonify({"ok": False, "error": "Error interno"}), 400


# ── Borrar empresa ─────────────────────────────────────────────────────────────

@app.route("/api/empresa/<int:eid>", methods=["DELETE"])
@login_required
def borrar_empresa(eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    try:
        # Obtener archivos de Storage antes de borrar
        facturas = sb.table("facturas_gastos").select("archivo_pdf").eq("empresa_id", eid).execute().data
        facturas += sb.table("facturas_venta").select("archivo_pdf").eq("empresa_id", eid).execute().data
        # Borrar registros en cascade
        sb.table("facturas_gastos").delete().eq("empresa_id", eid).execute()
        sb.table("facturas_venta").delete().eq("empresa_id", eid).execute()
        sb.table("empresas_clientes").delete().eq("id", eid).execute()
        # Borrar archivos de Storage
        import re as _re
        paths = []
        for f in facturas:
            url = f.get("archivo_pdf") or ""
            if url.startswith("http"):
                m = _re.search(r"/object/public/facturas/(.+)", url)
                if m:
                    paths.append(m.group(1))
        if paths:
            try:
                sb.storage.from_("facturas").remove(paths)
            except Exception:
                pass
        return jsonify({"ok": True})
    except Exception as ex:
        logging.exception('Error en endpoint')
        return jsonify({"ok": False, "error": "Error interno"}), 400


# ── Conciliación bancaria — cruce de extracto CSV vs facturas ─────────────────

@app.route("/api/empresa/<int:eid>/conciliacion", methods=["POST"])
@login_required
def conciliacion_bancaria(eid):
    err = validate_empresa_ownership(eid)
    if err: return err
    try:
        import pandas as pd
    except ImportError:
        return jsonify({"ok": False, "error": "pandas no disponible"}), 500
    import io

    if "archivo" not in request.files:
        return jsonify({"ok": False, "error": "No se recibió archivo"}), 400

    texto = request.files["archivo"].read().decode("utf-8", errors="replace")

    try:
        df = pd.read_csv(io.StringIO(texto))
    except Exception as ex:
        logging.exception("Error leyendo CSV")
        return jsonify({"ok": False, "error": "No se pudo leer el archivo CSV"}), 400

    df.columns = [c.strip().lower() for c in df.columns]

    monto_col = next((c for c in df.columns if any(k in c for k in ['monto','valor','importe','credito','credit','debit','debito'])), None)
    fecha_col = next((c for c in df.columns if any(k in c for k in ['fecha','date','dia'])), None)
    desc_col  = next((c for c in df.columns if any(k in c for k in ['descripcion','desc','concepto','referencia','detail'])), None)

    if not monto_col:
        cols = list(df.columns)
        return jsonify({"ok": False, "error": f"No se encontró columna de monto. Columnas detectadas: {cols}"}), 400

    df['_monto'] = (
        df[monto_col].astype(str)
        .str.replace(r'[$\s]', '', regex=True)
        .str.replace(r'[.,](?=\d{3})', '', regex=True)
        .str.replace(',', '.', regex=False)
    )
    df['_monto'] = pd.to_numeric(df['_monto'], errors='coerce').fillna(0).abs()

    ventas_rows = (
        sb.table("facturas_venta")
        .select("numero,valor_neto,cliente_nombre,fecha_vencimiento,estado")
        .eq("empresa_id", eid)
        .not_.eq("estado", "PAGADA")
        .execute().data
    )
    gastos_rows = (
        sb.table("facturas_gastos")
        .select("numero,valor_neto,proveedor_nombre,fecha_vencimiento,estado")
        .eq("empresa_id", eid)
        .not_.eq("estado", "PAGADA")
        .execute().data
    )
    # Normalise gastos to share campo 'cliente_nombre'
    for r in gastos_rows:
        r["cliente_nombre"] = r.pop("proveedor_nombre", "")

    todas = ventas_rows + gastos_rows
    coincidencias, sin_match = [], []

    for _, row in df.iterrows():
        monto = row['_monto']
        if monto < 1000:
            continue
        fecha = str(row[fecha_col]) if fecha_col else ''
        desc  = str(row[desc_col])  if desc_col  else ''
        match = None
        for f in todas:
            neto = f['valor_neto'] or 0
            if abs(monto - neto) <= max(neto * 0.015, 5000):
                match = f
                break
        if match:
            coincidencias.append({
                "extracto_monto":  int(monto),
                "extracto_fecha":  fecha,
                "extracto_desc":   desc[:60],
                "factura_numero":  match['numero'],
                "factura_tercero": match['cliente_nombre'],
                "factura_neto":    int(match['valor_neto'] or 0),
                "diferencia":      int(abs(monto - (match['valor_neto'] or 0))),
                "tipo":            "exacto" if abs(monto - (match['valor_neto'] or 0)) < 1000 else "aproximado",
            })
        else:
            sin_match.append({"monto": int(monto), "fecha": fecha, "descripcion": desc[:60]})

    return jsonify({
        "ok": True,
        "resumen": {
            "total_filas": len(df),
            "matches":     len(coincidencias),
            "sin_match":   len(sin_match),
            "pct_match":   round(len(coincidencias) / max(len(df), 1) * 100, 1),
        },
        "coincidencias": coincidencias,
        "sin_match":     sin_match[:30],
    })


# ── Calendario tributario ─────────────────────────────────────────────────────

@app.route("/api/calendario")
@login_required
def get_calendario():
    from datetime import date as d
    empresas = sb.table("empresas_clientes").select("id,nit,razon_social,ciudad,regimen").in_("id", get_user_empresa_ids() or [-1]).execute().data

    # Cargar obligaciones ya completadas
    try:
        comp_rows = sb.table("obligaciones_completadas").select("empresa_id,tipo,vencimiento").execute().data
        completadas = {(c["empresa_id"], c["tipo"], c["vencimiento"]) for c in comp_rows}
    except Exception:
        completadas = set()

    resultado = []
    hoy = d.today()
    for e in empresas:
        regimen = e.get("regimen") or "Juridica"
        obs = todas_las_obligaciones(e.get("nit", ""), regimen=regimen)
        for ob in obs:
            vto = d.fromisoformat(ob["vencimiento"])
            dias = (vto - hoy).days
            completada = (e["id"], ob["tipo"], ob["vencimiento"]) in completadas
            if completada:
                estado = "completada"
            elif dias < 0:
                estado = "vencida"
            elif dias <= 7:
                estado = "urgente"
            elif dias <= 30:
                estado = "proxima"
            else:
                estado = "ok"
            resultado.append({
                "empresa_id":     e["id"],
                "empresa":        e["razon_social"],
                "ciudad":         e.get("ciudad", ""),
                "nit":            e.get("nit", ""),
                "tipo":           ob["tipo"],
                "periodo":        ob["periodo"],
                "vencimiento":    ob["vencimiento"],
                "frecuencia":     ob["frecuencia"],
                "dias_restantes": dias,
                "estado":         estado,
                "completada":     completada,
            })
    resultado.sort(key=lambda x: x["vencimiento"])
    return jsonify({"ok": True, "obligaciones": resultado})


@app.route("/api/obligacion/completar", methods=["POST"])
@login_required
def completar_obligacion():
    body = request.get_json() or {}
    empresa_id  = body.get("empresa_id")
    tipo        = body.get("tipo")
    vencimiento = body.get("vencimiento")
    if not all([empresa_id, tipo, vencimiento]):
        return jsonify({"ok": False, "error": "Faltan campos"}), 400
    err = validate_empresa_ownership(empresa_id)
    if err: return err
    try:
        sb.table("obligaciones_completadas").upsert({
            "empresa_id": empresa_id,
            "tipo":        tipo,
            "vencimiento": vencimiento,
        }, on_conflict="empresa_id,tipo,vencimiento").execute()
        return jsonify({"ok": True})
    except Exception as e:
        logging.exception('Error en endpoint')
        return jsonify({"ok": False, "error": "Error interno"}), 500


@app.route("/api/obligacion/completar", methods=["DELETE"])
@login_required
def descompletar_obligacion():
    body = request.get_json() or {}
    empresa_id  = body.get("empresa_id")
    tipo        = body.get("tipo")
    vencimiento = body.get("vencimiento")
    if not all([empresa_id, tipo, vencimiento]):
        return jsonify({"ok": False, "error": "Faltan campos"}), 400
    err = validate_empresa_ownership(empresa_id)
    if err: return err
    try:
        sb.table("obligaciones_completadas").delete().eq("empresa_id", empresa_id).eq("tipo", tipo).eq("vencimiento", vencimiento).execute()
        return jsonify({"ok": True})
    except Exception as e:
        logging.exception('Error en endpoint')
        return jsonify({"ok": False, "error": "Error interno"}), 500


@app.route("/api/pendientes", methods=["GET"])
@login_required
def listar_pendientes():
    try:
        cid = session["contador_id"]
        rows = sb.table("empresas_pendientes").select("*").eq("contador_id", cid).order("created_at", desc=True).execute().data
        empresas = sb.table("empresas_clientes").select("id,nit,razon_social").eq("contador_id", cid).execute().data
        return jsonify({"ok": True, "pendientes": rows, "empresas": empresas})
    except Exception as e:
        logging.exception('Error en endpoint')
        return jsonify({"ok": False, "error": "Error interno"}), 500


@app.route("/api/pendientes/<pendiente_id>/asignar", methods=["POST"])
@login_required
def asignar_pendiente(pendiente_id):
    from datetime import datetime as dt
    body = request.get_json() or {}
    empresa_id = body.get("empresa_id")
    if not empresa_id:
        return jsonify({"ok": False, "error": "empresa_id requerido"}), 400

    err = validate_empresa_ownership(empresa_id)
    if err: return err

    try:
        cid = session["contador_id"]
        row = sb.table("empresas_pendientes").select("*").eq("id", pendiente_id).eq("contador_id", cid).execute().data
        if not row:
            return jsonify({"ok": False, "error": "No encontrado"}), 404
        p = row[0]
        datos = p.get("factura_data") or {}
        emp_row = sb.table("empresas_clientes").select("nit").eq("id", empresa_id).execute().data
        empresa_nit = emp_row[0]["nit"] if emp_row else ""
        if datos.get("numero"):
            guardar_factura(datos, int(empresa_id), empresa_nit, "", p.get("fuente", "upload"), sb)
        sb.table("empresas_pendientes").delete().eq("id", pendiente_id).execute()
        return jsonify({"ok": True})
    except Exception as e:
        logging.exception('Error en endpoint')
        return jsonify({"ok": False, "error": "Error interno"}), 500


@app.route("/api/pendientes/<pendiente_id>", methods=["DELETE"])
@login_required
def eliminar_pendiente(pendiente_id):
    try:
        cid = session["contador_id"]
        sb.table("empresas_pendientes").delete().eq("id", pendiente_id).eq("contador_id", cid).execute()
        return jsonify({"ok": True})
    except Exception as e:
        logging.exception('Error en endpoint')
        return jsonify({"ok": False, "error": "Error interno"}), 500


# ── Gmail OAuth multi-cliente ────────────────────────────────────────────────

def _google_creds():
    """Lee client_id y client_secret desde env vars o gmail_credentials.json."""
    client_id = os.environ.get("GOOGLE_CLIENT_ID", "")
    client_secret = os.environ.get("GOOGLE_CLIENT_SECRET", "")
    if not client_id or not client_secret:
        creds_path = Path(BASE_DIR) / "data" / "gmail_credentials.json"
        if creds_path.exists():
            raw = json.loads(creds_path.read_text())
            w = raw.get("web", raw.get("installed", {}))
            client_id = w.get("client_id", "")
            client_secret = w.get("client_secret", "")
    return client_id, client_secret

@app.route("/auth/gmail")
@login_required
def auth_gmail():
    empresa_id = request.args.get("empresa_id", "")
    email_hint = request.args.get("email", "")
    if not empresa_id:
        return "empresa_id requerido", 400
    client_id, _ = _google_creds()
    if not client_id:
        return "GOOGLE_CLIENT_ID no configurado en variables de entorno", 500
    redirect_uri = request.host_url.rstrip("/") + "/auth/gmail/callback"
    params = {
        "client_id": client_id,
        "redirect_uri": redirect_uri,
        "response_type": "code",
        "scope": "https://www.googleapis.com/auth/gmail.modify",
        "access_type": "offline",
        "prompt": "consent",
        "state": empresa_id,
    }
    if email_hint:
        params["login_hint"] = email_hint
    return redirect("https://accounts.google.com/o/oauth2/auth?" + urlencode(params))

@app.route("/auth/gmail/callback")
@login_required
def auth_gmail_callback():
    code = request.args.get("code", "")
    empresa_id = request.args.get("state", "")
    error = request.args.get("error", "")
    if error:
        return f"Google rechazó la autorización: {error}", 400
    if not code or not empresa_id:
        return "Parámetros inválidos", 400
    err = validate_empresa_ownership(int(empresa_id))
    if err: return err
    client_id, client_secret = _google_creds()
    redirect_uri = request.host_url.rstrip("/") + "/auth/gmail/callback"
    data = urlencode({
        "code": code,
        "client_id": client_id,
        "client_secret": client_secret,
        "redirect_uri": redirect_uri,
        "grant_type": "authorization_code",
    }).encode()
    try:
        req = _urllib_req.Request(
            "https://oauth2.googleapis.com/token", data=data, method="POST"
        )
        req.add_header("Content-Type", "application/x-www-form-urlencoded")
        with _urllib_req.urlopen(req, timeout=10) as resp:
            token_data = json.loads(resp.read())
    except Exception as e:
        return f"Error obteniendo tokens de Google: {e}", 500
    refresh_token = token_data.get("refresh_token", "")
    if not refresh_token:
        return "Google no envió refresh_token. Revoca el acceso en myaccount.google.com y vuelve a autorizar.", 400
    # Obtener email de la cuenta autorizada
    email = ""
    try:
        access_token = token_data.get("access_token", "")
        req2 = _urllib_req.Request("https://www.googleapis.com/gmail/v1/users/me/profile")
        req2.add_header("Authorization", f"Bearer {access_token}")
        with _urllib_req.urlopen(req2, timeout=5) as r:
            profile = json.loads(r.read())
            email = profile.get("emailAddress", "")
    except Exception:
        pass
    sb.table("gmail_tokens").upsert({
        "empresa_id": int(empresa_id),
        "email": email,
        "refresh_token": _encrypt_token(refresh_token),
        "token_created_at": datetime.now(timezone.utc).isoformat(),
        "activo": True,
    }, on_conflict="empresa_id").execute()
    # Activar Push Notifications (Pub/Sub watch) para esta cuenta
    try:
        from gmail_facturas import get_gmail_from_supabase, registrar_watch
        service_w, _ = get_gmail_from_supabase(int(empresa_id))
        if service_w:
            registrar_watch(service_w, int(empresa_id))
            logging.info("[gmail] Watch Pub/Sub registrado para empresa %s (%s)", empresa_id, email)
    except Exception:
        logging.exception("[gmail] No se pudo registrar watch para empresa %s", empresa_id)
    empresa = sb.table("empresas_clientes").select("razon_social,contador_id").eq("id", empresa_id).execute().data
    nombre      = empresa[0]["razon_social"] if empresa else f"Empresa {empresa_id}"
    contador_id = empresa[0].get("contador_id") if empresa else None
    # Notificar al contador por Telegram
    chat_id = _tg_chat_id_for_contador(contador_id)
    _tg_send_raw(chat_id,
        f"✅ *Gmail renovado correctamente*\n\n"
        f"🏢 Empresa: *{nombre}*\n"
        f"📧 Correo: {email}\n\n"
        f"El acceso quedó activo por 7 días más.\n"
        f"ContaBot seguirá leyendo las facturas automáticamente."
    )
    return f"""<html><body style="font-family:sans-serif;text-align:center;padding:60px;background:#f8fafc">
    <div style="max-width:440px;margin:0 auto;background:white;border-radius:16px;padding:2.5rem;box-shadow:0 4px 24px rgba(0,0,0,.08)">
    <div style="font-size:48px;margin-bottom:1rem">✅</div>
    <h2 style="color:#1e293b;margin-bottom:.5rem">Gmail renovado</h2>
    <p style="color:#475569;margin-bottom:.25rem"><b style="color:#1e293b">{email}</b></p>
    <p style="color:#475569;margin-bottom:1.5rem">conectado a <b style="color:#1e293b">{nombre}</b></p>
    <p style="background:#f0fdf4;border-radius:8px;padding:.75rem;color:#065f46;font-size:14px;margin-bottom:1.5rem">
    🔔 Se envió confirmación a tu Telegram</p>
    <a href="/" style="background:#2563eb;color:white;padding:12px 28px;border-radius:10px;text-decoration:none;font-weight:700;font-size:15px">
    Volver al dashboard →</a>
    </div></body></html>"""

@app.route("/api/gmail/tokens", methods=["GET"])
@login_required
def listar_gmail_tokens():
    empresa_ids = get_user_empresa_ids()
    if not empresa_ids:
        return jsonify({"ok": True, "tokens": []})
    tokens = sb.table("gmail_tokens").select("id,empresa_id,email,token_created_at,activo").in_("empresa_id", empresa_ids).execute().data
    return jsonify({"ok": True, "tokens": tokens})

@app.route("/api/gmail/tokens/<int:empresa_id>", methods=["DELETE"])
@login_required
def desconectar_gmail(empresa_id):
    err = validate_empresa_ownership(empresa_id)
    if err: return err
    sb.table("gmail_tokens").delete().eq("empresa_id", empresa_id).execute()
    return jsonify({"ok": True})

@app.route("/api/gmail/push", methods=["POST"])
def gmail_push():
    """Webhook Pub/Sub — Google avisa cuando llega email nuevo a una cuenta autorizada."""
    auth_header = request.headers.get("Authorization", "")
    if not auth_header.startswith("Bearer "):
        return "Unauthorized", 403
    token = auth_header[7:]
    try:
        from google.oauth2 import id_token as _id_token
        import google.auth.transport.requests as _ga_req
        _id_token.verify_oauth2_token(token, _ga_req.Request(), audience=request.base_url)
    except Exception:
        logging.warning("[push] Token JWT de Google inválido — acceso denegado")
        return "Unauthorized", 403
    import base64
    data    = request.get_json(silent=True) or {}
    message = data.get("message", {})
    if not message:
        return "ok", 200
    try:
        payload    = json.loads(base64.b64decode(message.get("data", "")).decode())
        email      = payload.get("emailAddress", "")
        history_id = str(payload.get("historyId", ""))
    except Exception:
        return "ok", 200
    if not email or not history_id:
        return "ok", 200
    token_row = sb.table("gmail_tokens").select("*").eq("email", email).eq("activo", True).execute().data
    if not token_row:
        return "ok", 200
    t          = token_row[0]
    empresa_id = t["empresa_id"]
    old_hid    = t.get("history_id", "")

    def _procesar():
        try:
            from gmail_facturas import get_gmail_from_supabase, escanear_desde_history
            service, _ = get_gmail_from_supabase(empresa_id)
            if service and old_hid:
                escanear_desde_history(service, empresa_id, email, old_hid)
            elif service:
                from gmail_facturas import escanear_inbox
                escanear_inbox(empresa_id=empresa_id, max_correos=20, service=service)
        except Exception:
            logging.exception("[push] Error procesando notificación %s", email)

    import threading
    threading.Thread(target=_procesar, daemon=True).start()
    return "ok", 200

@app.route("/api/gmail/activar-push", methods=["POST"])
@login_required
def activar_push_gmail():
    """Registra el watch de Pub/Sub para todas las cuentas activas."""
    try:
        from gmail_facturas import renovar_todos_los_watches
        renovar_todos_los_watches()
        return jsonify({"ok": True, "mensaje": "Watches registrados"})
    except Exception as e:
        logging.exception('Error en endpoint')
        return jsonify({"ok": False, "error": "Error interno"}), 500


@app.route("/api/empresa/<int:eid>/gmail/escanear", methods=["POST"])
@login_required
def escanear_gmail_empresa(eid):
    """Escaneo manual del inbox de Gmail para una empresa."""
    err = validate_empresa_ownership(eid)
    if err: return err
    body = request.get_json() or {}
    max_correos = min(int(body.get("max_correos", 50)), 2000)
    after_date = body.get("after_date")  # formato YYYY/MM/DD
    try:
        from gmail_facturas import get_gmail_from_supabase, escanear_inbox
        from google.auth.exceptions import RefreshError
        service, _ = get_gmail_from_supabase(eid)
        if not service:
            return jsonify({"ok": False, "error": "Esta empresa no tiene Gmail conectado"}), 400
        # Escaneos grandes corren en background para evitar timeout de Railway
        if max_correos > 100:
            cid = session["contador_id"]
            def _bg():
                try:
                    stats = escanear_inbox(empresa_id=eid, max_correos=max_correos, service=service, after_date=after_date)
                    _cache_invalidar(cid)
                    chat_id = _tg_chat_id_for_contador(cid)
                    _tg_send_raw(chat_id,
                        f"✅ *Escaneo Gmail completado*\n\n"
                        f"Empresa #{eid}\n"
                        f"Nuevas: {stats.get('nuevas',0)} | Duplicadas: {stats.get('duplicadas',0)} | "
                        f"Ignoradas: {stats.get('ignoradas',0)} | Errores: {stats.get('errores',0)}"
                    )
                except Exception:
                    logging.exception("Error en escaneo background empresa %s", eid)
            threading.Thread(target=_bg, daemon=True).start()
            return jsonify({"ok": True, "background": True,
                "mensaje": f"Escaneando {max_correos} correos en segundo plano. Te avisamos por Telegram cuando termine."})
        stats = escanear_inbox(empresa_id=eid, max_correos=max_correos, service=service, after_date=after_date)
        _cache_invalidar(session["contador_id"])
        return jsonify({"ok": True, **stats})
    except RefreshError:
        logging.warning("Token Gmail revocado para empresa %s", eid)
        try:
            token_row = sb.table("gmail_tokens").update({"activo": False}).eq("empresa_id", eid).execute()
            email_gmail = ""
            token_data = sb.table("gmail_tokens").select("email").eq("empresa_id", eid).execute().data
            if token_data:
                email_gmail = token_data[0].get("email", "")
            empresa = sb.table("empresas_clientes").select("razon_social,nit").eq("id", eid).execute().data
            nombre = empresa[0]["razon_social"] if empresa else f"empresa #{eid}"
            nit = empresa[0].get("nit", "") if empresa else ""
            cid = session.get("contador_id")
            chat_id = _tg_chat_id_for_contador(cid)
            _tg_send_raw(chat_id,
                f"⚠️ *Token Gmail vencido*\n\n"
                f"*Empresa:* {nombre}\n"
                f"*NIT:* {nit}\n"
                f"*Gmail:* {email_gmail}\n\n"
                f"Entra a ContaBot → selecciona *{nombre}* → tab *Gmail* → botón *Reconectar Gmail*"
            )
        except Exception:
            logging.exception("Error al manejar token revocado empresa %s", eid)
        return jsonify({"ok": False, "error": "El token de Gmail venció. Reconecta Gmail desde el tab Gmail de esta empresa."}), 401
    except Exception:
        logging.exception("Error escaneando Gmail empresa %s", eid)
        return jsonify({"ok": False, "error": "Error interno"}), 500


@app.route("/api/empresa/<int:eid>/facturas/limpiar", methods=["POST"])
@login_required
def limpiar_facturas_empresa(eid):
    """Borra TODAS las facturas de una empresa (solo para testing)."""
    err = validate_empresa_ownership(eid)
    if err: return err
    body = request.get_json() or {}
    confirmar = body.get("confirmar", False)
    if not confirmar:
        return jsonify({"ok": False, "error": "Envía confirmar: true para proceder"}), 400
    try:
        r_g = sb.table("facturas_gastos").delete().eq("empresa_id", eid).execute()
        r_v = sb.table("facturas_venta").delete().eq("empresa_id", eid).execute()
        _cache_invalidar(session["contador_id"])
        return jsonify({"ok": True, "gastos_borrados": len(r_g.data), "ventas_borradas": len(r_v.data)})
    except Exception:
        logging.exception("Error limpiando facturas empresa %s", eid)
        return jsonify({"ok": False, "error": "Error interno"}), 500


@app.route("/api/calendario/notificar", methods=["POST"])
@login_required
def notificar_obligaciones():
    """Envía Telegram con obligaciones que vencen en los próximos 7 días."""
    cid      = session["contador_id"]
    empresas = sb.table("empresas_clientes").select("id,nit,razon_social").eq("contador_id", cid).execute().data
    proximas = obligaciones_proximas(empresas, dias=7)
    if not proximas:
        return jsonify({"ok": True, "enviadas": 0})

    token   = os.environ.get("TELEGRAM_BOT_TOKEN", "")
    chat_id = _tg_chat_id_for_contador(cid)
    if not token or not chat_id:
        return jsonify({"ok": False, "error": "Telegram no configurado — vincula tu cuenta con /vincular en @contabot_contador_bot"})

    import urllib.request, urllib.parse, json as _json
    lineas = ["*Obligaciones tributarias próximas (7 días)*\n"]
    for p in proximas:
        ob   = p["obligacion"]
        dias = p["dias_restantes"]
        emoji = "🔴" if dias <= 2 else "🟡"
        lineas.append(
            f"{emoji} *{ob['tipo']}* — {p['empresa']}\n"
            f"   Período: {ob['periodo']} | Vence: {ob['vencimiento']} ({dias}d)"
        )
    texto = "\n".join(lineas)
    payload = _json.dumps({"chat_id": chat_id, "text": texto, "parse_mode": "Markdown"}).encode()
    req = urllib.request.Request(
        f"https://api.telegram.org/bot{token}/sendMessage",
        data=payload, headers={"Content-Type": "application/json"}
    )
    try:
        urllib.request.urlopen(req, timeout=5)
    except Exception as ex:
        return jsonify({"ok": False, "error": "Error interno"})
    return jsonify({"ok": True, "enviadas": len(proximas)})


# ── Subir factura electrónica DIAN (PDF / XML / ZIP) ─────────────────────────

@app.route("/api/subir-factura", methods=["POST"])
@login_required
@limiter.limit("20 per minute")
def subir_factura():
    from datetime import datetime as dt
    archivo = request.files.get("archivo")
    if not archivo:
        return jsonify({"ok": False, "error": "No se recibió archivo"}), 400

    fname = Path(archivo.filename or "").name  # solo basename, evita path traversal
    ext   = fname.lower().rsplit(".", 1)[-1] if "." in fname else ""
    if ext not in ("pdf", "xml", "zip"):
        return jsonify({"ok": False, "error": "Solo se aceptan PDF, XML o ZIP"}), 400

    data = archivo.read()
    tmp_dir = FACTURAS_DIR / "tmp"
    tmp_dir.mkdir(parents=True, exist_ok=True)

    # Descomprimir ZIP si aplica
    if ext == "zip":
        archivos = descomprimir_zip(data, tmp_dir)
        if not archivos:
            return jsonify({"ok": False, "error": "ZIP sin PDF/XML válido"}), 400
    else:
        file_path = tmp_dir / fname
        file_path.write_bytes(data)
        archivos = [file_path]

    # Extraer datos — modo híbrido: XML para estructura, PDF para completar huecos
    datos_xml, datos_pdf, file_usado = None, None, None
    for fp in archivos:
        if fp.suffix.lower() == ".xml":
            datos_xml = extraer_xml(str(fp))
            if datos_xml:
                file_usado = fp
        elif fp.suffix.lower() == ".pdf":
            datos_pdf = extraer_pdf(str(fp))
            if datos_pdf and not file_usado:
                file_usado = fp

    if datos_xml:
        datos = datos_xml
        if datos_pdf:
            # Completar campos que el XML dejó vacíos o en 0
            for campo in ["proveedor_nit", "proveedor_nombre", "receptor_nit", "receptor_nombre",
                          "total_factura", "valor_neto", "subtotal", "iva"]:
                if not datos.get(campo) and datos_pdf.get(campo):
                    datos[campo] = datos_pdf[campo]
    elif datos_pdf:
        datos = datos_pdf
    else:
        datos = None

    if not datos:
        return jsonify({"ok": False, "error": "No se pudieron extraer datos de la factura"}), 400

    # Detectar empresa por NIT receptor — solo entre las empresas del contador logueado
    cid = session["contador_id"]
    empresa = detectar_o_crear_empresa(datos, sb, contador_id=cid)

    # Fallback: si el XML dio NIT receptor pero no coincidió, intentar con el NIT del PDF
    if not empresa and datos_pdf and datos_pdf.get("receptor_nit") and datos_pdf.get("receptor_nit") != datos.get("receptor_nit"):
        datos["receptor_nit"] = datos_pdf["receptor_nit"]
        if datos_pdf.get("receptor_nombre"):
            datos["receptor_nombre"] = datos_pdf["receptor_nombre"]
        empresa = detectar_o_crear_empresa(datos, sb, contador_id=cid)

    empresa_id = empresa["id"] if empresa else None
    empresa_nombre = empresa["razon_social"] if empresa else "Empresa desconocida"

    # Empresa no reconocida — mostrar selector manual en el UI (no crear automáticamente)
    if not empresa_id:
        pendiente_id = guardar_empresa_pendiente(datos, fuente="upload", sb=sb, contador_id=cid)
        empresas_all = sb.table("empresas_clientes").select("id,nit,razon_social").eq("contador_id", cid).execute().data
        notificar_empresa_desconocida(datos, fuente="upload", pendiente_id=pendiente_id, empresas=empresas_all, sb=sb, contador_id=cid)
        return jsonify({
            "ok": True,
            "datos": datos,
            "empresa_detectada": False,
            "pendiente_id": pendiente_id,
            "empresas_disponibles": empresas_all,
            "mensaje": f"No se reconoció la empresa receptora (NIT: {datos.get('receptor_nit') or 'no encontrado'}). Seleccioná la empresa manualmente o créala primero.",
        })

    # Mover archivo a carpeta definitiva
    empresa_dir = FACTURAS_DIR / str(empresa_id)
    empresa_dir.mkdir(parents=True, exist_ok=True)
    destino = empresa_dir / (file_usado.name)
    file_usado.rename(destino)

    # Verificar duplicado
    numero = datos.get("numero", "")
    ya = sb.table("facturas_gastos").select("id").eq("empresa_id", empresa_id).eq("numero", numero).execute()
    if ya.data:
        return jsonify({
            "ok": True,
            "datos": datos,
            "empresa": empresa_nombre,
            "duplicada": True,
            "mensaje": f"La factura {numero} ya estaba registrada.",
        })

    fecha = datos.get("fecha") or dt.today().strftime("%Y-%m-%d")
    # Subir a Supabase Storage
    from extractor import subir_a_storage
    storage_url = subir_a_storage(str(destino), empresa_id, numero, fecha, sb)

    # Guardar en la tabla correcta según rol de la empresa (venta o gasto)
    flujo, _ = guardar_factura(datos, empresa_id, empresa["nit"],
                               storage_url or str(destino), "upload", sb)
    notificar_factura(datos, empresa_nombre, tipo=flujo, fuente="upload", sb=sb, contador_id=cid)

    return jsonify({
        "ok": True,
        "datos": datos,
        "empresa": empresa_nombre,
        "empresa_id": empresa_id,
        "duplicada": False,
        "mensaje": f"Factura {numero} registrada correctamente en {empresa_nombre}.",
    })


# ── Guardar factura con empresa seleccionada manualmente ─────────────────────

@app.route("/api/subir-factura/confirmar", methods=["POST"])
@login_required
def confirmar_factura():
    from datetime import datetime as dt
    body       = request.get_json()
    datos      = body.get("datos", {})
    empresa_id = body.get("empresa_id")
    if not empresa_id:
        return jsonify({"ok": False, "error": "empresa_id requerido"}), 400

    err = validate_empresa_ownership(empresa_id)
    if err: return err

    empresa_rows = sb.table("empresas_clientes").select("razon_social").eq("id", empresa_id).execute().data
    empresa_nombre = empresa_rows[0]["razon_social"] if empresa_rows else f"Empresa {empresa_id}"

    empresa_row = sb.table("empresas_clientes").select("nit").eq("id", empresa_id).execute().data
    empresa_nit = empresa_row[0]["nit"] if empresa_row else ""
    flujo, resultado = guardar_factura(datos, int(empresa_id), empresa_nit, "", "upload", sb)
    if resultado == "duplicada":
        return jsonify({"ok": True, "duplicada": True, "mensaje": f"La factura {datos.get('numero','')} ya estaba registrada."})
    notificar_factura(datos, empresa_nombre, tipo=flujo, fuente="upload", sb=sb, contador_id=session["contador_id"])
    return jsonify({"ok": True, "duplicada": False, "empresa": empresa_nombre,
                    "mensaje": f"Factura {datos.get('numero','')} registrada en {empresa_nombre} como {flujo}."})


# ── Telegram — helpers ───────────────────────────────────────────────────────

def _tg_chat_id_for_contador(contador_id) -> str:
    """Retorna telegram_chat_id del contador; fallback a la variable global."""
    if contador_id:
        try:
            rows = sb.table("contadores").select("telegram_chat_id").eq("id", contador_id).execute().data
            if rows and rows[0].get("telegram_chat_id"):
                return rows[0]["telegram_chat_id"]
        except Exception:
            pass
    return os.environ.get("TELEGRAM_CHAT_ID", "")


def _tg_send_raw(chat_id: str, texto: str):
    """Envía mensaje Telegram usando el token de entorno."""
    token = os.environ.get("TELEGRAM_BOT_TOKEN", "")
    if not token or not chat_id:
        return
    try:
        payload = json.dumps({"chat_id": chat_id, "text": texto, "parse_mode": "Markdown"}).encode()
        _urllib_req.urlopen(_urllib_req.Request(
            f"https://api.telegram.org/bot{token}/sendMessage",
            data=payload, headers={"Content-Type": "application/json"}
        ), timeout=5)
    except Exception as ex:
        logging.warning("[tg] No se pudo enviar: %s", ex)


# ── Telegram — vinculación por contador ──────────────────────────────────────

@app.route("/api/telegram/status")
@login_required
def telegram_status():
    cid = session["contador_id"]
    row = sb.table("contadores").select("telegram_chat_id").eq("id", cid).execute().data
    connected = bool(row and row[0].get("telegram_chat_id"))
    return jsonify({"connected": connected})


@app.route("/api/telegram/generar-link", methods=["POST"])
@login_required
def telegram_generar_link():
    cid   = session["contador_id"]
    token = secrets.token_urlsafe(16)
    sb.table("contadores").update({"telegram_token": token}).eq("id", cid).execute()
    bot_username = "contabot_contador_bot"
    link = f"https://t.me/{bot_username}?start={token}"
    return jsonify({"ok": True, "link": link})


@app.route("/api/telegram/desconectar", methods=["POST"])
@login_required
def telegram_desconectar():
    cid = session["contador_id"]
    sb.table("contadores").update({"telegram_chat_id": None, "telegram_token": None}).eq("id", cid).execute()
    return jsonify({"ok": True})


# ── Telegram Webhook ─────────────────────────────────────────────────────────

def _procesar_respuesta_empresa(chat_id: str, texto: str, token_bot: str):
    """
    Procesa la respuesta del usuario al teclado de asignación de empresa.
    Busca la factura pendiente más antigua para este contador, la asigna,
    y si hay más pendientes envía automáticamente el siguiente teclado.
    """
    contador_rows = sb.table("contadores").select("id").eq("telegram_chat_id", chat_id).execute().data
    if not contador_rows:
        return
    cid = contador_rows[0]["id"]

    pendientes = sb.table("empresas_pendientes").select("*").eq("contador_id", cid) \
        .order("created_at").limit(1).execute().data
    if not pendientes:
        return  # No hay nada pendiente — el usuario escribió algo random

    p       = pendientes[0]
    datos   = p.get("factura_data") or {}
    fuente  = p.get("fuente") or "gmail"
    num     = datos.get("numero", "—")

    if texto == "❌ No es de ningún cliente, ignorar":
        sb.table("empresas_pendientes").delete().eq("id", p["id"]).execute()
        _tg_send_raw(chat_id, f"🗑️ Factura N° {num} ignorada y eliminada de la bandeja.")
    else:
        emp = sb.table("empresas_clientes").select("id,razon_social,nit") \
            .eq("razon_social", texto).eq("contador_id", cid).execute().data
        if not emp:
            _tg_send_raw(chat_id, f"⚠️ No reconocí *{texto}* como cliente. Elige una opción del teclado.")
            return
        e = emp[0]
        flujo, _ = guardar_factura(datos, e["id"], e["nit"], "", fuente, sb)
        sb.table("empresas_pendientes").delete().eq("id", p["id"]).execute()
        tipo_label = "venta" if flujo == "venta" else "gasto"
        _tg_send_raw(chat_id,
            f"✅ *Factura asignada*\n\n"
            f"📌 Cliente: *{e['razon_social']}*\n"
            f"📄 Factura N° {num} — registrada como *{tipo_label}*"
        )

    # Siguiente en la cola
    resto = sb.table("empresas_pendientes").select("*").eq("contador_id", cid) \
        .order("created_at").limit(1).execute().data
    if resto:
        np = resto[0]
        ndatos  = np.get("factura_data") or {}
        nfuente = np.get("fuente") or "gmail"
        empresas_all = sb.table("empresas_clientes").select("razon_social") \
            .eq("contador_id", cid).execute().data
        from telegram_notif import _enviar_pregunta_empresa
        _enviar_pregunta_empresa(token_bot, chat_id, ndatos, nfuente, empresas_all)
    else:
        _tg_send_raw(chat_id, "✅ ¡Listo! No hay más facturas pendientes de asignar.")


@app.route("/api/telegram/webhook", methods=["POST"])
def telegram_webhook():
    """Recibe callbacks de Telegram y procesa confirmaciones de empresa vía reply_keyboard."""
    tg_secret = os.environ.get("TELEGRAM_WEBHOOK_SECRET", "")
    if not tg_secret:
        return "", 403
    received = request.headers.get("X-Telegram-Bot-Api-Secret-Token", "")
    if received != tg_secret:
        return "", 403
    update = request.get_json(silent=True) or {}
    token_bot = os.environ.get("TELEGRAM_BOT_TOKEN", "")

    # Manejar mensajes de texto
    if "message" in update:
        msg          = update["message"]
        text         = (msg.get("text") or "").strip()
        from_chat_id = str(msg["chat"]["id"])
        from telegram_notif import _send as _tg_send

        def _vincular(contador_row):
            """Guarda el chat_id en el contador y confirma."""
            sb.table("contadores").update({
                "telegram_chat_id": from_chat_id,
                "telegram_token":   None,
            }).eq("id", contador_row["id"]).execute()
            _tg_send(token_bot, from_chat_id,
                f"✅ ¡Hola {contador_row['nombre']}! Tu cuenta ContaBot quedó vinculada.\n\n"
                "🔔 Desde ahora recibirás aquí:\n"
                "• Facturas nuevas (correo automático y subidas manuales)\n"
                "• Aviso de empresa desconocida con selector de cliente directo\n"
                "• Alerta cuando el acceso Gmail esté por vencer\n"
                "• Recordatorios de obligaciones tributarias próximas\n\n"
                "Puedes responder a los botones directamente desde Telegram.")

        # /start — con o sin token de deep-link
        if text.startswith("/start"):
            parts      = text.split(None, 1)
            link_token = parts[1].strip() if len(parts) > 1 else ""
            if link_token:
                rows = sb.table("contadores").select("id,nombre").eq("telegram_token", link_token).execute().data
                if rows:
                    _vincular(rows[0])
                else:
                    _tg_send(token_bot, from_chat_id,
                        "⚠️ Enlace inválido o ya utilizado.\n"
                        "Usa /vincular tu@email.com para conectarte.")
            else:
                _tg_send(token_bot, from_chat_id,
                    "👋 Hola, soy ContaBot.\n\n"
                    "Para recibir notificaciones de tus clientes, envía:\n"
                    "  /vincular tu@email.com\n\n"
                    "(El email debe ser el que usaste para registrarte en ContaBot)")

        # /vincular email — forma directa sin necesitar el dashboard
        elif text.lower().startswith("/vincular"):
            parts = text.split(None, 1)
            email = parts[1].strip().lower() if len(parts) > 1 else ""
            if not email or "@" not in email:
                _tg_send(token_bot, from_chat_id,
                    "Por favor indica tu email:\n  /vincular tu@email.com")
            else:
                rows = sb.table("contadores").select("id,nombre").eq("email", email).execute().data
                if rows:
                    _vincular(rows[0])
                else:
                    _tg_send(token_bot, from_chat_id,
                        f"⚠️ No encontré ninguna cuenta con el email `{email}`.\n"
                        "Verifica que sea el mismo email con el que te registraste en ContaBot.")

        # Respuesta del teclado de opciones (reply_keyboard) para asignar empresa
        elif text and not text.startswith("/"):
            _procesar_respuesta_empresa(from_chat_id, text, token_bot)

        return jsonify({"ok": True})

    if "callback_query" not in update:
        return jsonify({"ok": True})

    cq         = update["callback_query"]
    cq_id      = cq["id"]
    data       = cq.get("data", "")
    chat_id    = str(cq["message"]["chat"]["id"])
    message_id = cq["message"]["message_id"]

    token = os.environ.get("TELEGRAM_BOT_TOKEN", "")

    def answer(text=""):
        """Cierra el spinner del botón."""
        try:
            import urllib.request as _ur
            payload = json.dumps({"callback_query_id": cq_id, "text": text}).encode()
            _ur.urlopen(_ur.Request(
                f"https://api.telegram.org/bot{token}/answerCallbackQuery",
                data=payload, headers={"Content-Type": "application/json"}
            ), timeout=5)
        except Exception:
            pass

    def edit_message(text):
        """Edita el mensaje original para mostrar el resultado."""
        try:
            import urllib.request as _ur
            payload = json.dumps({
                "chat_id": chat_id, "message_id": message_id,
                "text": text, "parse_mode": "Markdown",
            }).encode()
            _ur.urlopen(_ur.Request(
                f"https://api.telegram.org/bot{token}/editMessageText",
                data=payload, headers={"Content-Type": "application/json"}
            ), timeout=5)
        except Exception:
            pass

    if data.startswith("asignar_empresa:"):
        # Formato: asignar_empresa:pendiente_id:empresa_id
        partes = data.split(":")
        pendiente_id = partes[1] if len(partes) > 1 else None
        empresa_id   = int(partes[2]) if len(partes) > 2 else None
        answer("Asignando…")
        try:
            row = sb.table("empresas_pendientes").select("*").eq("id", pendiente_id).execute().data
            if not row:
                edit_message("⚠️ Esta factura ya fue procesada anteriormente.")
                return jsonify({"ok": True})
            p = row[0]
            datos  = p["factura_data"] or {}
            fuente = p["fuente"] or "upload"

            emp_row = sb.table("empresas_clientes").select("razon_social,nit").eq("id", empresa_id).execute().data
            empresa_nombre = emp_row[0]["razon_social"] if emp_row else f"Empresa {empresa_id}"
            empresa_nit    = emp_row[0]["nit"] if emp_row else ""

            numero = datos.get("numero", "")
            flujo = "gasto"
            if numero:
                flujo, _ = guardar_factura(datos, empresa_id, empresa_nit, "", fuente, sb)

            sb.table("empresas_pendientes").delete().eq("id", pendiente_id).execute()
            tipo_label = "venta" if flujo == "venta" else "gasto"
            edit_message(
                f"✅ *Factura asignada correctamente*\n\n"
                f"📌 Cliente: *{empresa_nombre}*\n"
                f"📄 Factura N° {numero or '—'} — registrada como *{tipo_label}*\n\n"
                f"Ya aparece en el dashboard de ContaBot."
            )
        except Exception as ex:
            logging.exception("[webhook] Error asignando empresa")
            edit_message("❌ Error interno.")

    elif data.startswith("confirmar_empresa:"):
        pendiente_id = data.split(":", 1)[1]
        answer("Procesando…")
        try:
            row = sb.table("empresas_pendientes").select("*").eq("id", pendiente_id).execute().data
            if not row:
                edit_message("⚠️ No se encontraron los datos pendientes. Puede que ya hayan sido procesados.")
                return jsonify({"ok": True})

            p = row[0]
            nit    = p["nit"]
            nombre = p["razon_social"] or f"Empresa NIT {nit}"
            ciudad = p["ciudad"] or ""
            datos  = p["factura_data"] or {}
            fuente = p["fuente"] or "upload"

            # Buscar si la empresa ya existe (por NIT, ignorando dígito verificador)
            from extractor import detectar_empresa as _detectar_empresa
            empresa_existente = _detectar_empresa(nit, sb)
            if empresa_existente:
                empresa_id = empresa_existente["id"]
                nombre     = empresa_existente["razon_social"]
            else:
                import random
                nueva = sb.table("empresas_clientes").insert({
                    "nit":         nit,
                    "razon_social": nombre,
                    "ciudad":      ciudad,
                    "sector":      "General",
                    "color":       random.choice(COLORES),
                    "icono":       random.choice(ICONOS),
                }).execute()
                empresa_id = nueva.data[0]["id"] if nueva.data else None

            if not empresa_id:
                edit_message("❌ Error creando la empresa. Inténtalo manualmente en ContaBot.")
                return jsonify({"ok": True})

            # Registrar la factura usando flujo correcto venta/gasto
            numero = datos.get("numero", "")
            flujo = "gasto"
            if numero:
                flujo, _ = guardar_factura(datos, empresa_id, nit, "", fuente, sb)

            sb.table("empresas_pendientes").delete().eq("id", pendiente_id).execute()
            tipo_label = "venta" if flujo == "venta" else "gasto"
            edit_message(
                f"✅ *Empresa creada y factura registrada*\n\n"
                f"🏢 *{nombre}*\n"
                f"🔢 NIT: `{nit}`\n"
                f"📄 Factura N° {numero or '—'} — registrada como *{tipo_label}*\n\n"
                f"Ya aparece en el dashboard de ContaBot."
            )
        except Exception as ex:
            logging.exception("[webhook] Error confirmando empresa")
            edit_message("❌ Error interno.")

    elif data.startswith("ignorar_empresa:"):
        pendiente_id = data.split(":", 1)[1]
        answer("Ignorado")
        try:
            row = sb.table("empresas_pendientes").select("nit,razon_social").eq("id", pendiente_id).execute().data
            nombre = row[0]["razon_social"] if row else "desconocida"
            nit    = row[0]["nit"] if row else "—"
            sb.table("empresas_pendientes").delete().eq("id", pendiente_id).execute()
            edit_message(f"🗑️ Factura de *{nombre}* (NIT `{nit}`) ignorada y eliminada.")
        except Exception as ex:
            logging.exception("[webhook] Error ignorando empresa")

    return jsonify({"ok": True})


# ── Cron jobs (APScheduler) ───────────────────────────────────────────────────

def _cron_gmail():
    """Renueva los watches de Pub/Sub (Gmail Push) cada 6 días."""
    try:
        sys.path.insert(0, SCRIPTS_DIR)
        from gmail_facturas import renovar_todos_los_watches
        renovar_todos_los_watches()
    except Exception as ex:
        logging.exception("[cron] Gmail watches error")

def _cron_tokens_gmail():
    """Avisa por Telegram al contador dueño de cada empresa cuando su token Gmail vence."""
    try:
        tokens = sb.table("gmail_tokens").select("*").eq("activo", True).execute().data
        host     = os.environ.get("RAILWAY_PUBLIC_DOMAIN", "")
        base_url = f"https://{host}" if host else "http://localhost:5000"
        for t in tokens:
            created_raw = t.get("token_created_at", "")
            if not created_raw:
                continue
            created = datetime.fromisoformat(created_raw.replace("Z", "+00:00"))
            dias = (datetime.now(timezone.utc) - created).days
            if dias < 6:
                continue
            # Obtener empresa y su contador_id
            emp = sb.table("empresas_clientes").select("razon_social,contador_id").eq("id", t["empresa_id"]).execute().data
            nombre      = emp[0]["razon_social"] if emp else f"Empresa {t['empresa_id']}"
            contador_id = emp[0].get("contador_id") if emp else None
            chat_id     = _tg_chat_id_for_contador(contador_id)
            link = f"{base_url}/auth/gmail?empresa_id={t['empresa_id']}&email={t['email']}"
            texto = (
                f"⚠️ *Token Gmail por vencer — {nombre}*\n"
                f"📧 {t['email']}\n\n"
                f"El acceso a Gmail vence mañana (día {dias}/7).\n"
                f"Toca el link para renovar (solo 30 segundos):\n"
                f"🔗 {link}\n\n"
                f"Solo haz clic en _Permitir_ — permisos ya preconfigurados."
            )
            _tg_send_raw(chat_id, texto)
            logging.info(f"[tokens] Aviso enviado a contador {contador_id}: {nombre} ({t['email']}) día {dias}")
    except Exception as ex:
        logging.exception("[tokens] Error chequeando tokens")

def _cron_obligaciones():
    """Notifica por Telegram al contador de cada empresa las obligaciones que vencen en 7 días."""
    try:
        # Agrupar empresas por contador_id
        todas = sb.table("empresas_clientes").select("id,nit,razon_social,contador_id").execute().data
        por_contador: dict = {}
        for e in todas:
            cid = e.get("contador_id")
            por_contador.setdefault(cid, []).append(e)

        for cid, empresas in por_contador.items():
            proximas = obligaciones_proximas(empresas, dias=7)
            if not proximas:
                continue
            chat_id = _tg_chat_id_for_contador(cid)
            if not chat_id:
                continue
            lineas = ["*Obligaciones tributarias — próximos 7 días*\n"]
            for p in proximas:
                ob    = p["obligacion"]
                dias  = p["dias_restantes"]
                emoji = "🔴" if dias <= 2 else "🟡"
                lineas.append(f"{emoji} *{ob['tipo']}* — {p['empresa']}\n   {ob['periodo']} | Vence {ob['vencimiento']} ({dias}d)")
            _tg_send_raw(chat_id, "\n".join(lineas))
    except Exception as ex:
        logging.exception("[cron] Obligaciones error")

def _iniciar_scheduler():
    try:
        from apscheduler.schedulers.background import BackgroundScheduler
        scheduler = BackgroundScheduler()
        scheduler.add_job(_cron_gmail,        "interval", days=6,        id="gmail_watches")   # renueva watches Push
        scheduler.add_job(_cron_obligaciones, "interval", hours=24,      id="obligaciones")
        scheduler.add_job(_cron_tokens_gmail, "cron",     hour=14, minute=0, id="tokens_gmail")  # 9am Colombia (UTC-5)
        scheduler.start()
        logging.info("[cron] Scheduler iniciado: Gmail Push (watches c/6d), tokens 9am, obligaciones c/24h")
    except ImportError:
        logging.warning("[cron] APScheduler no instalado — cron desactivado")
    except Exception as ex:
        logging.exception("[cron] Error iniciando scheduler")


def _migrar_tablas():
    """Crea tablas necesarias si no existen."""
    tablas = [
        ("empresas_pendientes",
         "CREATE TABLE IF NOT EXISTS empresas_pendientes ("
         "  id UUID DEFAULT gen_random_uuid() PRIMARY KEY,"
         "  nit TEXT, razon_social TEXT, ciudad TEXT,"
         "  factura_data JSONB, fuente TEXT,"
         "  created_at TIMESTAMPTZ DEFAULT NOW()"
         ");"),
        ("obligaciones_completadas",
         "CREATE TABLE IF NOT EXISTS obligaciones_completadas ("
         "  id UUID DEFAULT gen_random_uuid() PRIMARY KEY,"
         "  empresa_id INTEGER,"
         "  tipo TEXT,"
         "  vencimiento DATE,"
         "  realizada_en DATE DEFAULT CURRENT_DATE,"
         "  created_at TIMESTAMPTZ DEFAULT NOW(),"
         "  UNIQUE(empresa_id, tipo, vencimiento)"
         ");"),
        ("gmail_tokens",
         "CREATE TABLE IF NOT EXISTS gmail_tokens ("
         "  id SERIAL PRIMARY KEY,"
         "  empresa_id INTEGER REFERENCES empresas_clientes(id) ON DELETE CASCADE,"
         "  email TEXT NOT NULL,"
         "  refresh_token TEXT NOT NULL,"
         "  token_created_at TIMESTAMPTZ DEFAULT NOW(),"
         "  activo BOOLEAN DEFAULT TRUE,"
         "  UNIQUE(empresa_id)"
         ");"),
    ]
    for nombre, sql in tablas:
        try:
            sb.table(nombre).select("id").limit(1).execute()
        except Exception:
            try:
                sb.rpc("exec_sql", {"sql": sql}).execute()
                logging.info(f"[migración] Tabla {nombre} creada.")
            except Exception as ex:
                logging.error(f"[migración] {nombre} no existe y no se pudo crear: {ex}")
                logging.error(f"[migración] Crea manualmente: {sql}")


if __name__ == "__main__":
    port    = int(os.environ.get("PORT", 5000))
    is_local = port == 5000
    if is_local:
        import webbrowser, threading
        print("\n  ContaBot Demo — http://localhost:5000\n")
        threading.Timer(1.2, lambda: webbrowser.open("http://localhost:5000")).start()
    _migrar_tablas()
    _iniciar_scheduler()
    app.run(debug=False, host="0.0.0.0", port=port)
